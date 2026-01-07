import numbers
import os
import re
import time
from arrow import get
import numpy as np
import openpyxl
from openpyxl.styles import Border, Side, numbers
from openpyxl.utils import get_column_letter
import pandas as pd
import json
import threading
import queue
import uuid
import traceback
import urllib
import shutil
import zipfile
import logging

from flask import Flask, abort, g, request, render_template, redirect, url_for, flash, session, jsonify, send_file, has_request_context
from flask_bcrypt import Bcrypt
from datetime import datetime, timedelta
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename
from io import BytesIO
from dateutil import parser
from devtools import debug
from collections import defaultdict

from config.database import get_db_connection
from functions.popup_notification import render_alert

app = Flask(__name__)
app.secret_key = 'supersecretkey'
bcrypt = Bcrypt(app)

# Configure session timeout to 10 minutes
app.permanent_session_lifetime = timedelta(minutes=30)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Temporary storage for DataFrame
uploaded_data = None
uploaded_data_2 = None
uploaded_data_3 = None
uploaded_data_4 = None
uploaded_data_5 = None
uploaded_data_6 = None
uploaded_data_7 = None
uploaded_data_8 = None
uploaded_data_9 = None
uploaded_data_10 = None
uploaded_data_11 = None

data_available = False
flag = ''

active_facility_1 = None
active_facility_2 = None
active_facility_3 = None
active_facility_4 = None
active_facility_5 = None

closed_facility_1 = None
closed_facility_2 = None
closed_facility_3 = None
closed_facility_4 = None
closed_facility_5 = None

ALLOWED_EXTENSIONS = {'.txt'}
# Available flags
FLAGS = ["Individual", "Perusahaan"]
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
MAX_FILE_BIG_SIZE = 200 * 1024 * 1024  # 200MB
UPLOAD_BASE_DIR = "uploads"

# Create a global task queue and task results dictionary
task_queue = queue.Queue()
task_results = {}
task_progress = {}

# Thread-safe locks
task_progress_lock = threading.Lock()
task_id_registry_lock = threading.Lock()

# Task ID registry to prevent reuse
task_id_registry = {}  # Format: {task_id: {'created_at': timestamp, 'username': str, 'filename': str, 'status': str}}
active_uploads = defaultdict(set)  # Format: {username: set(task_ids)}

served_final_status = {}
conn = get_db_connection()
cur = conn.cursor()

def cleanup_old_tasks():
    """Remove old completed tasks from memory with proper validation"""
    
    def cleanup_worker():
        while True:
            try:
                current_time = time.time()
                
                # Configuration
                CLEANUP_DELAY_COMPLETED = 300  # 5 minutes after completion
                CLEANUP_DELAY_ERROR = 300      # 5 minutes after error
                CLEANUP_DELAY_STALE = 3600     # 1 hour for stale tasks
                REGISTRY_RETENTION = 7200      # 2 hours in registry before removal
                
                # Clean up task_progress
                tasks_to_remove = []
                
                with task_progress_lock:
                    for task_id, progress in list(task_progress.items()):
                        try:
                            status = progress.get('status', 'processing')
                            progress_value = progress.get('progress', 0)
                            timestamp = progress.get('timestamp', current_time)
                            
                            # Check if task should be cleaned
                            should_cleanup = False
                            reason = ""
                            
                            # Completed tasks - wait 5 minutes
                            if status == 'completed' and progress_value >= 100:
                                if current_time - timestamp > CLEANUP_DELAY_COMPLETED:
                                    should_cleanup = True
                                    reason = "completed and aged"
                            
                            # Error tasks - wait 5 minutes
                            elif status == 'error':
                                if current_time - timestamp > CLEANUP_DELAY_ERROR:
                                    should_cleanup = True
                                    reason = "error and aged"
                            
                            # Stale processing tasks (stuck) - wait 1 hour
                            elif status == 'processing':
                                if current_time - timestamp > CLEANUP_DELAY_STALE:
                                    should_cleanup = True
                                    reason = "stale/stuck processing"
                                    # Mark as error in registry
                                    mark_task_error(task_id)
                            
                            if should_cleanup:
                                tasks_to_remove.append((task_id, reason))
                        
                        except Exception as e:
                            app.logger.error(f"Error checking task {task_id}: {e}")
                
                # Remove tasks from memory
                for task_id, reason in tasks_to_remove:
                    with task_progress_lock:
                        removed_progress = task_progress.pop(task_id, None)
                    
                    with task_progress_lock:
                        task_results.pop(task_id, None)
                    
                    if removed_progress:
                        app.logger.info(f"🧹 Cleaned task {task_id} from memory (reason: {reason})")
                
                # Clean up old entries from registry (keep for 2 hours for audit)
                registry_to_remove = []
                
                with task_id_registry_lock:
                    for task_id, info in list(task_id_registry.items()):
                        created_at = info.get('created_at', current_time)
                        status = info.get('status', 'active')
                        
                        # Remove old completed/error entries from registry
                        if status in ['completed', 'error']:
                            completion_time = info.get('completed_at') or info.get('error_at', created_at)
                            if current_time - completion_time > REGISTRY_RETENTION:
                                registry_to_remove.append(task_id)
                        
                        # Remove very old active entries (shouldn't happen, but safety net)
                        elif status == 'active':
                            if current_time - created_at > CLEANUP_DELAY_STALE:
                                registry_to_remove.append(task_id)
                
                # Remove from registry
                for task_id in registry_to_remove:
                    with task_id_registry_lock:
                        removed_info = task_id_registry.pop(task_id, None)
                        
                        # Clean up from active_uploads
                        if removed_info:
                            username = removed_info.get('username')
                            if username and task_id in active_uploads[username]:
                                active_uploads[username].discard(task_id)
                        
                        if removed_info:
                            app.logger.info(f"🧹 Cleaned task {task_id} from registry")
                
                # Log statistics
                with task_progress_lock:
                    progress_count = len(task_progress)
                
                with task_id_registry_lock:
                    registry_count = len(task_id_registry)
                    active_count = sum(1 for info in task_id_registry.values() if info['status'] == 'active')
                
                if tasks_to_remove or registry_to_remove:
                    app.logger.info(
                        f"📊 Cleanup stats - Progress: {progress_count}, "
                        f"Registry: {registry_count} (Active: {active_count}), "
                        f"Removed: {len(tasks_to_remove)} progress + {len(registry_to_remove)} registry"
                    )
                
                # Sleep for 1 minute before next cleanup
                time.sleep(60)
                
            except Exception as e:
                app.logger.error(f"❌ Error in cleanup worker: {e}")
                import traceback
                app.logger.error(traceback.format_exc())
                time.sleep(60)  # Continue even on error
    
    # Start cleanup thread
    cleanup_thread = threading.Thread(target=cleanup_worker, daemon=True, name="TaskCleanupWorker")
    cleanup_thread.start()
    app.logger.info("✅ Task cleanup worker started")

def clean_data_for_sql(data):
    """Clean data before SQL insertion"""
    if isinstance(data, pd.DataFrame):
        # Replace NaN with None
        data = data.replace({np.nan: None, pd.NaT: None})
        
        # Convert problematic columns
        for col in data.columns:
            # Handle numeric columns
            if data[col].dtype in ['float64', 'float32']:
                # Round to reasonable precision (adjust as needed)
                data[col] = data[col].round(2)
                # Replace inf with None
                data[col] = data[col].replace([np.inf, -np.inf], None)
            
            # Clean string columns that might contain numeric data
            elif data[col].dtype == 'object':
                # Replace empty strings with None for potential numeric fields
                data[col] = data[col].replace('', None)
                data[col] = data[col].replace('nan', None)
    
    return data

def try_parse_upload_date(uploadDate):
    # Format: "29 July 2025, 16:26"
    try:
        return datetime.strptime(uploadDate, "%d %B %Y, %H:%M")
    except ValueError:
        pass

    # Format: "2025-07-29 17:41:57"
    try:
        return datetime.strptime(uploadDate, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        pass

    # Format dengan microseconds: "2025-12-04 17:06:16.627000"
    try:
        return datetime.strptime(uploadDate, "%Y-%m-%d %H:%M:%S.%f")
    except ValueError:
        pass

    # Format dari JS locale: "29/7/2025, 17.29.35"
    try:
        return datetime.strptime(uploadDate, "%d/%m/%Y, %H.%M.%S")
    except ValueError:
        pass

    raise ValueError(f"Format tanggal tidak dikenali: {uploadDate}")

# Tambahkan helper function
def is_allowed_file(filename, mimetype):
    """Check if file is allowed (.txt and correct MIME type)"""
    return (
        filename.lower().endswith('.txt') and 
        mimetype in ['text/plain', 'application/octet-stream']
    )
    
def is_txt_file(filename):
    return os.path.splitext(filename)[1].lower() in ALLOWED_EXTENSIONS

# --- Utility Functions ---
def sanitize_folder_name(nama_file):
    """Konversi nama file menjadi nama folder yang valid"""
    return nama_file.lower().replace(" ", "_").replace("-", "_")

def ensure_upload_dir():
    """Pastikan direktori upload utama ada"""
    if not os.path.exists(UPLOAD_BASE_DIR):
        os.makedirs(UPLOAD_BASE_DIR)

def create_upload_folder(nama_file_upload):
    """Buat folder untuk upload session tertentu"""
    ensure_upload_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    folder_name = sanitize_folder_name(nama_file_upload)
    folder_name = f"{folder_name}_{timestamp}"
    upload_path = os.path.join(UPLOAD_BASE_DIR, folder_name)
    
    os.makedirs(upload_path, exist_ok=True)
    return upload_path

def save_file_to_local(content, filename, folder_path):
    """Save file content to local folder"""
    try:
        os.makedirs(folder_path, exist_ok=True)
        file_path = os.path.join(folder_path, filename)
        
        with open(file_path, 'wb') as f:
            f.write(content)
        
        app.logger.info(f"📁 Saved file: {file_path}")
        return file_path
    
    except Exception as e:
        app.logger.error(f"Error saving file {filename}: {e}")
        raise

# --- Check if file name exists ---
def check_filename_exists(namaFileUpload):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        query = "SELECT COUNT(*) FROM slik_uploader WHERE namaFileUpload = ?"
        cursor.execute(query, (namaFileUpload,))
        count = cursor.fetchone()[0]
        
        return count > 0
        
    except Exception as e:
        print(f"Error checking filename: {e}")
        traceback.print_exc()
        return False
    finally:
        cursor.close()
        conn.close()

def get_existing_filenames():
    """
    Fungsi untuk mengambil daftar nama file yang sudah ada
    Sesuaikan dengan struktur database/storage Anda
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT DISTINCT namaFileUpload FROM slik_uploader")
        result = cursor.fetchall()
        return [row[0] for row in result]
        
    except Exception as e:
        print(f"Error getting existing filenames: {e}")
        return []

# --- Alternative: Jika menggunakan Flask dengan flash message ---
def save_file_metadata_to_db_flask(periodeData, namaFileUpload, uploadFolderPath, username, fullname, roleAccess, uploadDate):
    try:
        
        conn = get_db_connection()
        cursor = conn.cursor()

        query = """
            INSERT INTO slik_uploader (periodeData, namaFileUpload, uploadFolderPath, username, fullname, roleAccess, uploadDate)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """
        cursor.execute(query, (periodeData, namaFileUpload, uploadFolderPath, username, fullname, roleAccess, uploadDate))
        conn.commit()
        
        return True

    except Exception as e:
        print(f"Error saving to DB: {e}")
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()

# --- Background Worker ---
def process_files_worker():
    """
    Background worker yang terintegrasi dengan:
    - Unique task ID system
    - Upload session tracking
    - Thread-safe progress updates
    - Proper cleanup and error handling
    """
    while True:
        try:
            task_id, files, user_info = task_queue.get(timeout=1)
        except queue.Empty:
            continue

        conn = None
        validation_failed = False
        upload_folder_path = None

        try:
            # ✅ PERBAIKAN 1: Validasi task masih aktif
            if not is_task_active(task_id):
                app.logger.warning(f"⚠️ Task {task_id} sudah tidak aktif, skip processing")
                task_queue.task_done()
                continue
            
            # ✅ PERBAIKAN 2: Log start processing dengan session info
            with task_id_registry_lock:
                task_info = task_id_registry.get(task_id, {})
                upload_session = task_info.get('upload_session')
            
            app.logger.info(
                f"🚀 Worker started - Task: {task_id}, Session: {upload_session}, "
                f"User: {user_info.get('username')}, File: {user_info.get('nama_file')}"
            )
            
            bulan_dict = {
                1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
                5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
                9: "September", 10: "Oktober", 11: "November", 12: "Desember"
            }

            nama_file_upload = user_info.get("nama_file")
            upload_folder_path = create_upload_folder(nama_file_upload)

            periode_data = None
            uploaded_files = []

            session_npwp = None
            session_identitas = None

            # ✅ PERBAIKAN 3: Update progress dengan helper function
            update_task_progress(task_id, 5, 'processing', 'Memvalidasi file...')

            # Prepare uploaded files
            for file_data in files:
                filename = secure_filename(file_data['filename'])
                mimetype = file_data.get('mimetype', 'text/plain')
                
                if not is_allowed_file(filename, mimetype):
                    error_msg = f"File {filename} bukan .txt atau tipe MIME salah."
                    
                    # ✅ PERBAIKAN: Gunakan update_task_progress
                    update_task_progress(task_id, 0, 'error', error_msg)
                    
                    task_results[task_id] = {
                        'status': 'error',
                        'message': error_msg,
                        'error_type': 'validation_error',
                        'completed': True,
                        'redirect_url': '/upload-big-size'
                    }
                    
                    # Cleanup dan return
                    if upload_folder_path and os.path.exists(upload_folder_path):
                        shutil.rmtree(upload_folder_path)
                    
                    mark_task_error(task_id)
                    task_queue.task_done()
                    return

                save_file_to_local(file_data['content'], filename, upload_folder_path)

                uploaded_files.append(FileStorage(
                    stream=BytesIO(file_data['content']),
                    filename=filename,
                    name='file',
                    content_type=mimetype
                ))

            # ✅ PERBAIKAN 4: Update progress per validation stage
            update_task_progress(task_id, 8, 'processing', f'Memvalidasi {len(uploaded_files)} file...')

            # Loop untuk validasi semua file dulu
            for idx, uploaded_file in enumerate(uploaded_files):
                # ✅ Update progress per file
                file_progress = 8 + int(7 * ((idx + 1) / len(uploaded_files)))
                update_task_progress(
                    task_id, 
                    file_progress, 
                    'processing', 
                    f'Memvalidasi file {idx + 1} dari {len(uploaded_files)}...'
                )
                
                file_content = uploaded_file.stream.read()
                file_processed = False

                for encoding in ['utf-8', 'latin-1', 'utf-16', 'ascii']:
                    try:
                        content = re.sub(r'[^\x20-\x7E\t\r\n]', '', file_content.decode(encoding))
                        data = json.loads(content)

                        if 'perusahaan' in data:
                            obj = data['perusahaan']
                            ident_key = 'npwp'
                        elif 'individual' in data:
                            obj = data['individual']
                            ident_key = 'noIdentitas'
                        else:
                            continue

                        posisi = obj.get('posisiDataTerakhir', '')
                        ident = obj.get('parameterPencarian', {}).get(ident_key, '')

                        if not posisi or not ident:
                            error_msg = f"File {uploaded_file.filename} tidak memiliki data posisi atau identitas."
                            
                            # ✅ PERBAIKAN: Gunakan update_task_progress
                            update_task_progress(task_id, 0, 'error', error_msg)
                            
                            task_results[task_id] = {
                                'status': 'error',
                                'message': error_msg,
                                'error_type': 'validation_error',
                                'completed': True,
                                'redirect_url': '/upload-big-size'
                            }
                            
                            if os.path.exists(upload_folder_path):
                                shutil.rmtree(upload_folder_path)
                            
                            mark_task_error(task_id)
                            validation_failed = True
                            break

                        # Validasi NPWP consistency
                        if ident_key == 'npwp':
                            if session_npwp is None:
                                session_npwp = ident
                            elif session_npwp != ident:
                                error_msg = f"NPWP file ke-{idx+1} ({ident}) tidak konsisten dengan file pertama ({session_npwp})."
                                
                                update_task_progress(task_id, 0, 'error', error_msg)
                                
                                task_results[task_id] = {
                                    'status': 'error',
                                    'message': error_msg,
                                    'error_type': 'validation_error',
                                    'completed': True,
                                    'redirect_url': '/upload-big-size'
                                }
                                
                                if os.path.exists(upload_folder_path):
                                    shutil.rmtree(upload_folder_path)
                                
                                mark_task_error(task_id)
                                validation_failed = True
                                break
                        
                        # Validasi Identitas consistency
                        else:
                            if session_identitas is None:
                                session_identitas = ident
                            elif session_identitas != ident:
                                error_msg = f"Identitas file ke-{idx+1} ({ident}) tidak konsisten dengan file pertama ({session_identitas})."
                                
                                update_task_progress(task_id, 0, 'error', error_msg)
                                
                                task_results[task_id] = {
                                    'status': 'error',
                                    'message': error_msg,
                                    'error_type': 'validation_error',
                                    'completed': True,
                                    'redirect_url': '/upload-big-size'
                                }
                                
                                if os.path.exists(upload_folder_path):
                                    shutil.rmtree(upload_folder_path)
                                
                                mark_task_error(task_id)
                                validation_failed = True
                                break

                        # Parse periode data dari posisi
                        date_obj = datetime.strptime(posisi, "%Y%m")
                        periode_data = f"{bulan_dict[date_obj.month]} {date_obj.year}"
                        
                        # ✅ PERBAIKAN: Update metadata dengan thread-safe
                        with task_progress_lock:
                            if task_id in task_progress and 'temp_metadata' in task_progress[task_id]:
                                task_progress[task_id]['temp_metadata']['periodeData'] = periode_data

                        file_processed = True
                        break

                    except (json.JSONDecodeError, ValueError, Exception) as e:
                        continue

                # Reset stream position untuk file berikutnya
                uploaded_file.stream.seek(0)
                
                if validation_failed:
                    break

                # Jika tidak ada encoding yang berhasil
                if not file_processed:
                    error_msg = f"Gagal memproses file: {uploaded_file.filename}"
                    
                    update_task_progress(task_id, 0, 'error', error_msg)
                    
                    task_results[task_id] = {
                        "status": "error",
                        "message": error_msg,
                        "error_type": "validation_error",
                        "completed": True,
                        "redirect_url": "/upload-big-size"
                    }
                    
                    app.logger.error(f"[{task_id}] {error_msg}")
                    
                    if os.path.exists(upload_folder_path):
                        shutil.rmtree(upload_folder_path)
                    
                    mark_task_error(task_id)
                    validation_failed = True
                    break
                
            if validation_failed:
                task_queue.task_done()
                return
            
            # ✅ PERBAIKAN 5: Validasi selesai
            update_task_progress(task_id, 15, 'processing', 'Validasi selesai, menyimpan metadata...')
            
            # Insert metadata 1x
            uploaded_at = datetime.now()
            
            if periode_data:
                # ✅ PERBAIKAN: Update metadata dengan format yang konsisten
                metadata_update = {
                    'periodeData': periode_data,
                    'uploadDate': uploaded_at.strftime('%d %B %Y, %H:%M'),
                    'namaFileUpload': nama_file_upload,
                    'username': user_info.get("username"),
                    'fullname': user_info.get("fullname")
                }
                
                with task_progress_lock:
                    if task_id in task_progress:
                        if 'temp_metadata' not in task_progress[task_id]:
                            task_progress[task_id]['temp_metadata'] = {}
                        task_progress[task_id]['temp_metadata'].update(metadata_update)
                
                # Save to database
                save_file_metadata_to_db_flask(
                    periodeData=periode_data,
                    namaFileUpload=nama_file_upload,
                    uploadFolderPath=upload_folder_path,
                    username=user_info.get("username"),
                    fullname=user_info.get("fullname"),
                    roleAccess=user_info.get("role_access"),
                    uploadDate=uploaded_at
                )

            # ✅ PERBAIKAN 6: Process files dengan progress tracking yang sudah diperbaiki
            update_task_progress(task_id, 20, 'processing', 'Memproses file...')
            
            result = process_uploaded_files(task_id, files, uploaded_files, user_info, uploaded_at)

            # ✅ PERBAIKAN 7: Handle result dengan proper cleanup
            if isinstance(result, dict) and result.get("error"):
                error_msg = result.get("message", "Terjadi kesalahan.")
                error_type = result.get("error_type", "validation_error")
                
                # Update progress menggunakan helper
                update_task_progress(task_id, 0, 'error', error_msg)
                
                # Store error result
                task_results[task_id] = {
                    "status": "error",
                    "message": error_msg,
                    "error_type": error_type,
                    "completed": True,
                    "redirect_url": result.get("redirect_url", "/upload-big-size")
                }

                # Cleanup folder
                if os.path.exists(upload_folder_path):
                    shutil.rmtree(upload_folder_path)
                    app.logger.info(f"🧹 Cleaned up folder: {upload_folder_path}")

                mark_task_error(task_id)
                app.logger.error(f"❌ [{task_id}] Failed: {error_msg}")

            else:
                # ✅ PERBAIKAN 8: Success case dengan metadata final
                result['upload_folder_path'] = upload_folder_path
                
                # Prepare final metadata
                final_metadata = {
                    'periodeData': periode_data,
                    'username': user_info.get("username"),
                    'namaFileUpload': nama_file_upload,
                    'uploadDate': uploaded_at.strftime('%d %B %Y, %H:%M'),
                    'fullname': user_info.get("fullname")
                }
                
                # ✅ PERBAIKAN: Update progress to completed dengan metadata
                success = update_task_progress(
                    task_id, 
                    100, 
                    'completed', 
                    'Upload berhasil diproses',
                    metadata=final_metadata
                )
                
                if not success:
                    app.logger.warning(f"⚠️ Failed to update progress for completed task {task_id}")
                
                # Store success result
                task_results[task_id] = {
                    "status": "success",
                    "result": result,
                    "completed": True,
                    "redirect_url": "/upload-success",
                    "message": "File berhasil diproses"
                }
                
                # Mark task as completed in registry
                mark_task_completed(task_id)
                
                app.logger.info(
                    f"✅ [{task_id}] Completed successfully - "
                    f"Session: {upload_session}, File: {nama_file_upload}"
                )

        except Exception as e:
            error_msg = f"Error tidak terduga: {str(e)}"
            
            # ✅ PERBAIKAN: Gunakan update_task_progress untuk error
            update_task_progress(task_id, 0, 'error', error_msg)
            
            task_results[task_id] = {
                "status": "error",
                "message": error_msg,
                "error_type": "system_error",
                "completed": True,
                "redirect_url": "/upload-big-size"
            }
            
            mark_task_error(task_id)
            
            app.logger.error(f"❌ [{task_id}] Unexpected error: {error_msg}")
            app.logger.error(traceback.format_exc())
            
            # Cleanup folder on error
            if upload_folder_path and os.path.exists(upload_folder_path):
                try:
                    shutil.rmtree(upload_folder_path)
                    app.logger.info(f"🧹 Cleaned up folder after error: {upload_folder_path}")
                except Exception as cleanup_error:
                    app.logger.error(f"Error cleaning up folder: {cleanup_error}")

        finally:
            # ✅ PERBAIKAN 9: Proper cleanup
            if conn:
                try:
                    conn.close()
                except:
                    pass
            
            # Always mark task as done
            task_queue.task_done()
            
            # Log final status
            with task_id_registry_lock:
                if task_id in task_id_registry:
                    final_status = task_id_registry[task_id].get('status', 'unknown')
                    app.logger.info(f"🏁 [{task_id}] Worker finished - Final status: {final_status}")
                   
# ==================== WORKER INITIALIZATION ====================

# Initialize cleanup system
cleanup_old_tasks()

# Start worker threads
NUM_WORKERS = 4
worker_threads = []

for i in range(NUM_WORKERS):
    t = threading.Thread(
        target=process_files_worker, 
        daemon=True,
        name=f"FileProcessWorker-{i+1}"
    )
    t.start()
    worker_threads.append(t)
    app.logger.info(f"✅ Started worker thread: FileProcessWorker-{i+1}")

app.logger.info(f"✅ All {NUM_WORKERS} worker threads started successfully")

def escape_sql(val):
    """Mencegah SQL Injection dengan mengganti ' menjadi ''."""
    return val.replace("'", "''") if isinstance(val, str) else val

# --- Download Functions ---
def create_zip_from_folder(folder_path, zip_name=None):
    """Buat file ZIP dari semua file dalam folder"""
    if not os.path.exists(folder_path):
        raise FileNotFoundError(f"Folder {folder_path} tidak ditemukan")
    
    if zip_name is None:
        zip_name = f"{os.path.basename(folder_path)}.zip"
    
    # Buat ZIP file di memory
    memory_file = BytesIO()
    
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                # Tambahkan file ke ZIP dengan path relatif
                arcname = os.path.relpath(file_path, folder_path)
                zipf.write(file_path, arcname)
    
    memory_file.seek(0)
    return memory_file, zip_name

# Route untuk mengecek nama file
@app.route('/check-filename', methods=['POST'])
def check_filename():
    try:
        data = request.get_json()
        nama_file = data.get('nama_file', '').strip()
        
        if not nama_file:
            return jsonify({'exists': False, 'message': 'Nama file tidak boleh kosong'})
        
        # Cek apakah nama file sudah ada di database
        exists = check_filename_exists(nama_file)
        
        return jsonify({
            'exists': exists,
            'message': 'Nama file sudah ada' if exists else 'Nama file tersedia'
        })
        
    except Exception as e:
        print(f"Error checking filename: {e}")
        return jsonify({'exists': False, 'message': 'Error checking filename'}), 500

@app.route('/list_files')
def list_files():
    return render_template('list_file_uploads.html')

# --- Get Upload List ---
@app.route('/list_uploads')
def list_uploads():
    """List semua upload dengan info folder path dan tanggal"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Menggunakan uploadFolderPath sebagai pengganti fileContentBase64
        cursor.execute("""
            SELECT id, namaFileUpload, uploadFolderPath, uploadDate, username, periodeData
            FROM slik_uploader 
            ORDER BY uploadDate DESC
        """)
        
        uploads = []
        for row in cursor.fetchall():
            upload_id, nama_file_upload, folder_path, upload_date, username, periode_data = row
            
            # Check if folder still exists and count files
            folder_exists = os.path.exists(folder_path) if folder_path else False
            file_count = 0
            
            if folder_exists:
                try:
                    file_count = len([f for f in os.listdir(folder_path) 
                                    if os.path.isfile(os.path.join(folder_path, f))])
                except:
                    file_count = 0
            
            # Format tanggal untuk display yang lebih baik
            formatted_date = upload_date
            if upload_date:
                try:
                    # Jika upload_date adalah string, parse dulu
                    if isinstance(upload_date, str):
                        from datetime import datetime
                        date_obj = datetime.strptime(upload_date, '%Y-%m-%d %H:%M:%S')
                        formatted_date = date_obj.isoformat()
                    else:
                        # Jika sudah datetime object
                        formatted_date = upload_date.isoformat()
                except:
                    formatted_date = str(upload_date)
            
            uploads.append({
                "id": upload_id,
                "nama_file": nama_file_upload,  # Sesuai dengan HTML template
                "folder_path": folder_path,
                "file_count": file_count,
                "upload_date": formatted_date,
                "username": username,
                "periode": periode_data,  # Sesuai dengan HTML template
                "folder_exists": folder_exists,
                "download_url": f"/download_upload/{upload_id}" if folder_exists else None
            })
            
        return jsonify(uploads)
        
    except Exception as e:
        print(f"Error listing uploads: {e}")
        return jsonify({"error": "Terjadi kesalahan"}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

# --- Flask Routes untuk Download ---
@app.route('/download_upload')
def download_upload_zip():
    """Download semua file dalam satu upload session sebagai ZIP"""
    periodeData = request.args.get('periodeData')
    username = session.get('username')
    namaFileUpload = request.args.get('namaFileUpload')
    uploadDate = request.args.get('uploadDate')

    downloadType = "File Upload TXT"
    downloadDate = datetime.now()
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # FIX: Query hanya memiliki 5 parameter, bukan 7
            query = """
                INSERT INTO slik_download_logging (periodeData, namaFileUpload, downloadType, username, downloadDate)
                VALUES (?, ?, ?, ?, ?)
            """
            
            cursor.execute(query, (periodeData, namaFileUpload, downloadType, username, downloadDate))
            conn.commit()

        except Exception as e:
            print(f"Error saving to DB: {e}")
            traceback.print_exc()
        
        # Ambil info upload dari database (menggunakan uploadFolderPath)
        cursor.execute("""
            SELECT uploadFolderPath, namaFileUpload 
            FROM slik_uploader 
            WHERE periodeData = ? AND namaFileUpload = ?
        """, (periodeData, namaFileUpload))
        
        result = cursor.fetchone()
        
        if not result:
            return jsonify({"error": "Upload tidak ditemukan"}), 404
        
        upload_folder_path, nama_file_upload = result
        
        # Pastikan folder masih ada
        if not os.path.exists(upload_folder_path):
            return jsonify({"error": "File tidak ditemukan di server"}), 404
        
        # Buat ZIP file
        zip_file, zip_name = create_zip_from_folder(
            upload_folder_path, 
            f"{sanitize_folder_name(nama_file_upload)}.zip"
        )
        
        # Return ZIP file
        return send_file(
            zip_file,
            as_attachment=True,
            download_name=zip_name,
            mimetype='application/zip'
        )
        
    except Exception as e:
        print(f"Error downloading files: {e}")
        traceback.print_exc()
        return jsonify({"error": "Terjadi kesalahan saat mengunduh file"}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

@app.errorhandler(404)
def not_found_error(error):
    return jsonify({'success': False, 'message': 'Resource not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({'success': False, 'message': 'Internal server error'}), 500

def validate_password(password):
    """Validate password strength"""
    if len(password) < 9:
        return False, "Password must be at least 9 characters long"
    if not re.search(r'[A-Z]', password):
        return False, "Password must contain at least one uppercase letter"
    if not re.search(r'\d', password):
        return False, "Password must contain at least one number"
    if not re.search(r'[@$!%*?&]', password):
        return False, "Password must contain at least one special character (@$!%*?&)"
    return True, "Password is valid"

def validate_email(email):
    """Validate email format"""
    pattern = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    return re.match(pattern, email) is not None

@app.route('/user-management')
def user_management():
    """Render user management page"""
    username = session.get('username')
    fullname = session.get('fullname')
    role_access = session.get('role_access')
    report_access = session.get('report_access')
    return render_template(
        'user_management.html',
        username=username,
        fullname=fullname,
        role_access=role_access,
        report_access=report_access
        )
    
@app.route('/api/users', methods=['GET'])
def get_users():
    """Get all users"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        query = """
            SELECT id, username, fullname, email, role_access, report_access, created_date
            FROM [dbo].[users]
            ORDER BY created_date DESC
        """
        
        cursor.execute(query)
        rows = cursor.fetchall()
        
        users = []
        for row in rows:
            users.append({
                'id': row.id,
                'username': row.username,
                'fullname': row.fullname,
                'email': row.email,
                'role_access': row.role_access,
                'report_access': row.report_access,
                'created_date': row.created_date.isoformat() if row.created_date else None
            })
        
        cursor.close()
        conn.close()
        
        return jsonify(users)
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['GET'])
def get_user(user_id):
    """Get single user by ID"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        query = """
            SELECT id, username, fullname, email, role_access, report_access, created_date
            FROM [dbo].[users]
            WHERE id = ?
        """
        
        cursor.execute(query, (user_id,))
        row = cursor.fetchone()
        
        if not row:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        user = {
            'id': row.id,
            'username': row.username,
            'fullname': row.fullname,
            'email': row.email,
            'role_access': row.role_access,
            'report_access': row.report_access,
            'created_date': row.created_date.isoformat() if row.created_date else None
        }
        
        cursor.close()
        conn.close()
        
        return jsonify(user)
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/users', methods=['POST'])
def create_user():
    """Create new user"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['username', 'password', 'fullname', 'email', 'role_access', 'report_access']
        for field in required_fields:
            if field not in data or data[field] == '' or data[field] is None:
                return jsonify({'success': False, 'message': f'{field} is required'}), 400
        
        username = data['username']
        password = data['password']
        fullname = data['fullname']
        email = data['email']
        role_access = data['role_access']
        report_access = data['report_access']
        
        # Validate email
        if not validate_email(email):
            return jsonify({'success': False, 'message': 'Invalid email format'}), 400
        
        # Validate password
        is_valid, message = validate_password(password)
        if not is_valid:
            return jsonify({'success': False, 'message': message}), 400

        # Hash password
        password_hash = bcrypt.generate_password_hash(password)
        # Convert bytes to string if needed
        if isinstance(password_hash, bytes):
            password_hash = password_hash.decode('utf-8')

        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if username already exists
        cursor.execute("SELECT id FROM [dbo].[users] WHERE username = ?", (username,))
        if cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Username already exists'}), 400

        # Check if email already exists
        cursor.execute("SELECT id FROM [dbo].[users] WHERE email = ?", (email,))
        if cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Email already exists'}), 400

        # Insert new user
        query = """
            INSERT INTO [dbo].[users]
            (username, password_hash, fullname, email, role_access, report_access, created_date)
            VALUES (?, ?, ?, ?, ?, ?, GETDATE())
        """

        cursor.execute(query, (username, password_hash, fullname, email, role_access, report_access))
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'message': 'User created successfully'}), 201
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    """Update existing user"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['username', 'fullname', 'email', 'role_access', 'report_access']
        for field in required_fields:
            if field not in data or data[field] == '' or data[field] is None:
                return jsonify({'success': False, 'message': f'{field} is required'}), 400
        
        username = data['username']
        fullname = data['fullname']
        email = data['email']
        role_access = data['role_access']
        report_access = data['report_access']
        
        # Validate email
        if not validate_email(email):
            return jsonify({'success': False, 'message': 'Invalid email format'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if user exists
        cursor.execute("SELECT id FROM [dbo].[users] WHERE id = ?", (user_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Check if username already exists (excluding current user)
        cursor.execute("SELECT id FROM [dbo].[users] WHERE username = ? AND id != ?", (username, user_id))
        if cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Username already exists'}), 400
        
        # Check if email already exists (excluding current user)
        cursor.execute("SELECT id FROM [dbo].[users] WHERE email = ? AND id != ?", (email, user_id))
        if cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Email already exists'}), 400
        
        # Update user
        query = """
            UPDATE [dbo].[users]
            SET username = ?, fullname = ?, email = ?, role_access = ?, report_access = ?
            WHERE id = ?
        """
        
        cursor.execute(query, (username, fullname, email, role_access, report_access, user_id))
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'message': 'User updated successfully'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    """Delete user"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if user exists
        cursor.execute("SELECT id FROM [dbo].[users] WHERE id = ?", (user_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Delete user
        query = "DELETE FROM [dbo].[users] WHERE id = ?"
        cursor.execute(query, (user_id,))
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'message': 'User deleted successfully'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# Login Route
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        session.clear()  # Bersihkan session lama
        session['data_available'] = False

    elif request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # Connect ke database
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT password_hash, role_access, fullname, report_access
            FROM users
            WHERE username = ?
        """, (username,))
        user = cur.fetchone()

        if user and bcrypt.check_password_hash(user[0], password):
            fullname = user[2]
            role_access = user[1]
            report_access = user[3]

            # Simpan data ke session
            session.permanent = True
            session['username'] = username
            session['fullname'] = fullname
            session['role_access'] = role_access
            session['report_access'] = report_access
            session['upload_done'] = True

            print(f"User '{username}' login dengan report_access = {report_access}")

            # Redirect berdasarkan level akses
            if report_access == 0:
                return redirect(url_for('upload_big_size_file'))
            elif report_access == 1:
                return redirect(url_for('daftar_fasilitas_debitur_page'))
            else:
                # report_access == 2 atau nilai lainnya
                return redirect(url_for('upload_big_size_file'))

        else:
            flash("Invalid username or password.")
    
    return render_template('login.html')

@app.route('/api/users/<int:user_id>/change-password', methods=['PUT'])
def change_password(user_id):
    """Change user password"""
    try:
        data = request.get_json()
        
        if 'password' not in data or not data['password']:
            return jsonify({'success': False, 'message': 'Password is required'}), 400
        
        password = data['password']
        
        # Validate password
        is_valid, message = validate_password(password)
        if not is_valid:
            return jsonify({'success': False, 'message': message}), 400
        
        # Hash password
        password_hash = bcrypt.generate_password_hash(password)
        # Convert bytes to string if needed
        if isinstance(password_hash, bytes):
            password_hash = password_hash.decode('utf-8')

        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if user exists
        cursor.execute("SELECT id FROM [dbo].[users] WHERE id = ?", (user_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'User not found'}), 404

        # Update password
        query = "UPDATE [dbo].[users] SET password_hash = ? WHERE id = ?"
        cursor.execute(query, (password_hash, user_id))
        conn.commit()

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'message': 'Password changed successfully'})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/users/<int:user_id>/reset-password', methods=['POST'])
def reset_password_user(user_id):
    """Reset user password (admin functionality)"""
    try:
        data = request.get_json()

        if 'password' not in data or not data['password']:
            return jsonify({'success': False, 'message': 'Password is required'}), 400

        password = data['password']

        # Validate password
        is_valid, message = validate_password(password)
        if not is_valid:
            return jsonify({'success': False, 'message': message}), 400

        # Hash password
        password_hash = bcrypt.generate_password_hash(password)
        # Convert bytes to string if needed
        if isinstance(password_hash, bytes):
            password_hash = password_hash.decode('utf-8')

        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if user exists
        cursor.execute("SELECT id FROM [dbo].[users] WHERE id = ?", (user_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'User not found'}), 404

        # Update password
        query = "UPDATE [dbo].[users] SET password_hash = ? WHERE id = ?"
        cursor.execute(query, (password_hash, user_id))
        conn.commit()

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'message': 'Password reset successfully'})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# Logout Route
@app.route('/logout')
def logout():
    session.pop('username', None)
    flash("You have been logged out.")
    return redirect(url_for('login'))

def insert_data(cursor, table_name, data_item, columns_to_remove=None, extra_columns=None):
    """
    Fungsi umum untuk menyisipkan data ke tabel tertentu dengan opsi metadata tambahan.
    """
    if not data_item:
        return
    
    try:
        # Filter kolom jika perlu
        filtered_item = data_item
        if columns_to_remove:
            if isinstance(data_item, dict):
                filtered_item = {k: v for k, v in data_item.items() if k not in columns_to_remove}
        
        # Penyesuaian key tertentu (khusus untuk tabel tertentu)
        if table_name == "slik_ringkasan_fasilitas" and "krediturBPR/S" in filtered_item:
            filtered_item["krediturBPR_S"] = filtered_item.pop("krediturBPR/S")
        
        # Tambahkan kolom tambahan
        if extra_columns:
            filtered_item.update(extra_columns)
        
        # Siapkan query
        columns = ', '.join(filtered_item.keys())
        placeholders = ', '.join(['?'] * len(filtered_item))
        values = tuple(filtered_item.values())

        query = f"""
            INSERT INTO {table_name} ({columns})
            VALUES ({placeholders})
        """
        cursor.execute(query, values)

    except Exception as e:
        print(f"Error inserting data into {table_name}: {e}")
        traceback.print_exc()
        print(f"Item: {data_item}")

def generate_unique_task_id(username, filename):
    """
    Generate a truly unique task ID that includes:
    - Timestamp for uniqueness
    - Username for tracking
    - Random UUID component
    - Hash of filename
    """
    timestamp = int(time.time() * 1000)  # milliseconds
    random_component = str(uuid.uuid4())[:8]
    
    # Create a unique identifier
    task_id = f"{username}_{timestamp}_{random_component}"
    
    # Register the task ID
    with task_id_registry_lock:
        # Ensure absolutely unique
        while task_id in task_id_registry:
            random_component = str(uuid.uuid4())[:8]
            task_id = f"{username}_{timestamp}_{random_component}"
        
        # Register in registry
        task_id_registry[task_id] = {
            'created_at': time.time(),
            'username': username,
            'filename': filename,
            'status': 'active',
            'upload_session': timestamp  # Use this to identify unique upload sessions
        }
        
        # Track active uploads per user
        active_uploads[username].add(task_id)
    
    app.logger.info(f"✅ Generated unique task_id: {task_id} for user: {username}, file: {filename}")
    return task_id

def is_task_active(task_id):
    """Check if a task is currently active"""
    with task_id_registry_lock:
        if task_id not in task_id_registry:
            return False
        return task_id_registry[task_id]['status'] == 'active'

def mark_task_completed(task_id):
    """Mark a task as completed in the registry"""
    with task_id_registry_lock:
        if task_id in task_id_registry:
            task_id_registry[task_id]['status'] = 'completed'
            task_id_registry[task_id]['completed_at'] = time.time()
            
            # Remove from active uploads
            username = task_id_registry[task_id].get('username')
            if username and task_id in active_uploads[username]:
                active_uploads[username].discard(task_id)
            
            app.logger.info(f"✅ Marked task {task_id} as completed")

def mark_task_error(task_id):
    """Mark a task as errored in the registry"""
    with task_id_registry_lock:
        if task_id in task_id_registry:
            task_id_registry[task_id]['status'] = 'error'
            task_id_registry[task_id]['error_at'] = time.time()
            
            # Remove from active uploads
            username = task_id_registry[task_id].get('username')
            if username and task_id in active_uploads[username]:
                active_uploads[username].discard(task_id)
            
            app.logger.info(f"❌ Marked task {task_id} as error")

def update_task_progress(task_id, progress, status='processing', message=None, metadata=None):
    """Thread-safe progress update with validation"""
    
    # Validate task is active
    if not is_task_active(task_id) and status == 'processing':
        app.logger.warning(f"⚠️ Attempted to update inactive task: {task_id}")
        return False
    
    with task_progress_lock:
        # Initialize if not exists
        if task_id not in task_progress:
            task_progress[task_id] = {
                'created_at': time.time(),
                'upload_session': None
            }
        
        # Get upload session from registry
        with task_id_registry_lock:
            if task_id in task_id_registry:
                task_progress[task_id]['upload_session'] = task_id_registry[task_id]['upload_session']
        
        # Update progress
        task_progress[task_id]['progress'] = int(progress)
        task_progress[task_id]['status'] = status
        task_progress[task_id]['timestamp'] = time.time()
        task_progress[task_id]['last_update'] = datetime.now().isoformat()
        
        if message:
            task_progress[task_id]['message'] = message
        if metadata:
            task_progress[task_id]['temp_metadata'] = metadata
        
        # Update progress bars for detailed tracking
        if 'progress_bars' not in task_progress[task_id]:
            task_progress[task_id]['progress_bars'] = {}
        
        task_progress[task_id]['progress_bars']['file_processing'] = min(progress, 80)
        task_progress[task_id]['progress_bars']['db_processing'] = max(0, progress - 80)
        
        # Mark as completed/error in registry if done
        if status == 'completed':
            mark_task_completed(task_id)
        elif status == 'error':
            mark_task_error(task_id)
    
    return True

def process_uploaded_files(task_id, files, uploaded_files, user_info, uploaded_at):
    """
    Process uploaded files with task-scoped variables (not global).
    This prevents race conditions when multiple uploads happen simultaneously.
    """
    
    # ========== TASK-SCOPED VARIABLES (NOT GLOBAL) ==========
    # All variables are local to this task to prevent concurrent upload conflicts
    
    uploaded_data = None
    uploaded_data_2 = None
    uploaded_data_3 = None
    uploaded_data_4 = None
    uploaded_data_5 = None
    uploaded_data_6 = None
    uploaded_data_7 = None
    uploaded_data_8 = None
    uploaded_data_9 = None
    uploaded_data_10 = None
    uploaded_data_11 = None
    
    flag = ''

    active_facility_1 = None
    active_facility_2 = None
    active_facility_3 = None
    active_facility_4 = None
    active_facility_5 = None

    closed_facility_1 = None
    closed_facility_2 = None
    closed_facility_3 = None
    closed_facility_4 = None
    closed_facility_5 = None
    
    columns_to_remove = ['agunan', 'penjamin']
    
    all_uploaded_data = []
    list_debitur = []
    list_uploaded_data_6 = []
    list_uploaded_data_7 = []
    list_uploaded_data_8 = []
    list_uploaded_data_9 = []
    list_uploaded_data_10 = []
    
    table_data = None
    list_table_data = []
    conn = None

    json_header = None
    json_individual = None
    json_perusahaan = None
    json_paramPencarian = None
    json_dpdebitur = None
    json_kPengurusPemilik = None
    json_rFasilitas = None
    json_fKreditPembiayan = None
    json_fSuratBerharga = None
    json_fLC = None
    json_fGaransi = None
    json_fFasilitasLain = None

    df_kPengurusPemilik = None
    df_temp = None
    data_temp = None
    df_expanded = None

    # ========== CRITICAL: TASK-SCOPED IDENTITAS & NPWP ==========
    # These must be local to prevent cross-task contamination
    task_session_identitas = None  # Validasi identitas untuk task ini
    task_session_npwp = None       # Validasi NPWP untuk task ini
    
    jenis_surat_berharga = [
        {
            "Jenis Surat Berharga": "Sertifikat Bank Indonesia (SBI)",
            "Kode": "F0401"
        },
        {
            "Jenis Surat Berharga": "Sertifikat Deposito Bank Indonesia (SDBI)",
            "Kode": "F0403"
        },
        {
            "Jenis Surat Berharga": "Surat Berharga Bank Indonesia (SBBI) dalam",
            "Kode": "F0404"
        },
        {
            "Jenis Surat Berharga": "Surat Perbendaharaan Negara (SPN)",
            "Kode": "F040501"
        },
        {
            "Jenis Surat Berharga": "Surat Perbendaharaan Negara Syariah",
            "Kode": "F040502"
        },
        {
            "Jenis Surat Berharga": "(SIMA)",
            "Kode": "F0406"
        },
        {
            "Jenis Surat Berharga": "Promes/Aksep",
            "Kode": "F0408"
        },
        {
            "Jenis Surat Berharga": "Wesel - Wesel Ekspor",
            "Kode": "F040901"
        },
        {
            "Jenis Surat Berharga": "(SKBDN)",
            "Kode": "F040902"
        },
        {
            "Jenis Surat Berharga": "Wesel - Lainnya",
            "Kode": "F040999"
        },
        {
            "Jenis Surat Berharga": "Surat Berharga Komersial",
            "Kode": "F0410"
        },
        {
            "Jenis Surat Berharga": "Medium Term Notes (MTN)",
            "Kode": "F041101"
        },
        {
            "Jenis Surat Berharga": "Medium Term Notes (MTN) Syariah",
            "Kode": "F041102"
        },
        {
            "Jenis Surat Berharga": "Floating Rate Notes (FRN)",
            "Kode": "F0412"
        },
        {
            "Jenis Surat Berharga": "Credit Linked Notes",
            "Kode": "F0413"
        },
        {
            "Jenis Surat Berharga": "Reksadana",
            "Kode": "F041401"
        },
        {
            "Jenis Surat Berharga": "Reksadana Syariah",
            "Kode": "F041402"
        },
        {
            "Jenis Surat Berharga": "Reksadana Dana Pendapatan Tetap",
            "Kode": "F041403"
        },
        {
            "Jenis Surat Berharga": "Obligasi Dalam rangka program rekapitalisasi",
            "Kode": "F04150101"
        },
        {
            "Jenis Surat Berharga": "Obligasi Negara (ON)",
            "Kode": "F04150102"
        },
        {
            "Jenis Surat Berharga": "Obligasi Ritel Indonesia (ORI)",
            "Kode": "F04150103"
        },
        {
            "Jenis Surat Berharga": "Obligasi Korporasi - Subordinasi",
            "Kode": "F0415010501"
        },
        {
            "Jenis Surat Berharga": "Obligasi Korporasi - Non Subordinasi",
            "Kode": "F0415010602"
        },
        {
            "Jenis Surat Berharga": "Obligasi Lainnya",
            "Kode": "F04150199"
        },
        {
            "Jenis Surat Berharga": "Sukuk Bank Indonesia",
            "Kode": "F04150201"
        },
        {
            "Jenis Surat Berharga": "Sukuk Negara",
            "Kode": "F04150203"
        },
        {
            "Jenis Surat Berharga": "Sukuk Ritel",
            "Kode": "F04150204"
        },
        {
            "Jenis Surat Berharga": "Tjarah Fixed Rate",
            "Kode": "F04150205"
        },
        {
            "Jenis Surat Berharga": "Sukuk Korporasi - Subordinasi",
            "Kode": "F0415020601"
        },
        {
            "Jenis Surat Berharga": "Sukuk Korporasi - Non Subordinasi",
            "Kode": "F0415020602"
        },
        {
            "Jenis Surat Berharga": "Project Based Sukuk (PBS)",
            "Kode": "F04150207"
        },
        {
            "Jenis Surat Berharga": "Sukuk Valas Bank Indonesia (SUVBI)",
            "Kode": "F04150208"
        },
        {
            "Jenis Surat Berharga": "Sukuk Lainnya",
            "Kode": "F04150299"
        },
        {
            "Jenis Surat Berharga": "Dana Investasi Real Estate (DIRE)",
            "Kode": "F0416"
        },
        {
            "Jenis Surat Berharga": "Efek Beragun Aset",
            "Kode": "F041701"
        },
        {
            "Jenis Surat Berharga": "Efek Beragun Aset Syariah",
            "Kode": "F041702"
        },
        {
            "Jenis Surat Berharga": "Sekuritas Rupiah Bank Indonesia (SRBI)",
            "Kode": "F0422"
        },
        {
            "Jenis Surat Berharga": "Sekuritas Valas Bank Indonesia (SVBI)",
            "Kode": "F0423"
        },
        {
            "Jenis Surat Berharga": "Surat Berharga Lainnya",
            "Kode": "F0499"
        }
    ]

    username = user_info['username']
    nama_file = user_info['nama_file']
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    total_files = len(files)
    
    identitas = None
    
    npwp = None
    
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        conn.autocommit = False
        
        for idx, uploaded_file in enumerate(uploaded_files, start=1):
            progress = 10 + int(70 * (idx / total_files))
            if task_id in task_progress:
                task_progress[task_id]['progress'] = progress
            
            filename = uploaded_file.filename
            if not filename.lower().endswith('.txt'):
                # Set error status dan stop processing
                task_progress[task_id]['status'] = 'error'
                return {
                    "error": True,
                    "error_type": "validation_error",
                    "message": f"File {filename} must have a .txt extension",
                    "redirect_url": "/upload-big-size"
                }
                
            if uploaded_file.mimetype != 'text/plain':
                # Set error status dan stop processing
                task_progress[task_id]['status'] = 'error'
                return {
                    "error": True,
                    "error_type": "validation_error", 
                    "message": f"File {filename} must be a plain text (.txt) file",
                    "redirect_url": "/upload-big-size"
                }
            
            if uploaded_file:
                try:
                    encodings = ['utf-8', 'utf-16', 'latin-1', 'ascii']
                    content = None
                    file_content = uploaded_file.stream.read()
                    
                    for encoding in encodings:
                        try:
                            uncleaned_content = file_content.decode(encoding, errors='ignore').strip()
                            content = re.sub(r'[^\x20-\x7E\t\r\n]', '', uncleaned_content)
                            break  # Jika berhasil, keluar dari loop
                        except (UnicodeDecodeError, LookupError):
                            continue  # Jika gagal, coba encoding berikutnya
                    
                    if content is None or len(content) == 0:
                        task_progress[task_id]['status'] = 'error'
                        return {
                            "error": True,
                            "error_type": "validation_error",
                            "message": f"Could not decode file {filename} with supported encodings",
                            "redirect_url": "/upload-big-size"
                        }

                    # Parse JSON dari konten
                    try:
                        data = json.loads(content)
                    except json.JSONDecodeError as e:
                        # Set error status dan stop processing
                        task_progress[task_id]['status'] = 'error'
                        return {
                            "error": True,
                            "error_type": "validation_error",
                            "message": f"Invalid JSON file {filename}: {str(e)}",
                            "redirect_url": "/upload-big-size"
                        }
                    
                    bulan_dict = {
                        1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
                        5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
                        9: "September", 10: "Oktober", 11: "November", 12: "Desember"
                    }
                    
                    json_header = data['header']
                    if 'perusahaan' in data and isinstance(data['perusahaan'], dict) and data['perusahaan'].get('posisiDataTerakhir'):
                        data_perusahaan = data['perusahaan']
                        posisiDataTerakhir = data_perusahaan['posisiDataTerakhir']
                        json_paramPencarian = data_perusahaan.get('parameterPencarian', {})
                        npwp = json_paramPencarian.get('npwp', '')
                        app.logger.debug(f"[{task_id}] Get NPWP: {npwp}")
                        
                        # Validasi NPWP untuk sesi upload (task-scoped)
                        if task_session_npwp is None:
                            # File pertama - simpan NPWP sebagai referensi
                            task_session_npwp = npwp
                            app.logger.info(f"[{task_id}] Task session NPWP set to: {npwp}")
                        elif npwp != task_session_npwp:
                            error_msg = f"File yang diupload tidak konsisten. File pertama: {task_session_npwp}, File ke-{idx}: {npwp}. Pastikan semua file dalam satu sesi upload memiliki NPWP yang sama."
                            app.logger.warning(f"[{task_id}] {error_msg}")
                            return {
                                "error": True,
                                "error_type": "validation_error",
                                "message": error_msg,
                                "redirect_url": "/upload-big-size"
                            }
                    elif 'individual' in data and isinstance(data['individual'], dict) and data['individual'].get('posisiDataTerakhir'):
                        data_individual = data['individual']
                        posisiDataTerakhir = data_individual['posisiDataTerakhir']
                        json_paramPencarian = data_individual.get('parameterPencarian', {})
                        identitas = json_paramPencarian.get('noIdentitas', '')
                        
                        # Validasi identitas untuk sesi upload
                        if task_session_identitas is None:
                            # File pertama - simpan identitas sebagai referensi
                            task_session_identitas = identitas
                            app.logger.info(f"[{task_id}] Task session identitas set to: {identitas}")
                        elif identitas != task_session_identitas:
                            # File kedua dan seterusnya - bandingkan dengan identitas file pertama
                            error_msg = f"File yang diupload tidak konsisten. File pertama: {task_session_identitas}, File ke-{idx}: {identitas}. Pastikan semua file dalam satu sesi upload memiliki Nomor Identitas yang sama."
                            app.logger.warning(f"[{task_id}] {error_msg}")
                            return {
                                "error": True,
                                "error_type": "validation_error",
                                "message": error_msg,
                                "redirect_url": "/upload-big-size"
                            }
                                
                    date_obj = datetime.strptime(posisiDataTerakhir, "%Y%m")
                    json_posisiDataTerakhir = f"{bulan_dict[date_obj.month]} {date_obj.year}"
                    
                    if json_header:
                        try:
                            # Buat salinan tanpa key yang tidak dibutuhkan
                            filtered_item = {k: v for k, v in json_header.items() if k not in columns_to_remove}

                            # Siapkan nama kolom tambahan
                            base_columns = ['periodeData', 'username', 'namaFileUpload', 'uploadDate']
                            base_values = [json_posisiDataTerakhir, username, nama_file, current_datetime]

                            # Gabungkan kolom dan nilai tambahan dengan hasil filter
                            all_columns = base_columns + list(filtered_item.keys())
                            all_values = tuple(base_values + list(filtered_item.values()))

                            # Buat string kolom dan placeholders
                            columns = ', '.join(all_columns)
                            placeholders = ', '.join(['?'] * len(all_columns))
                            

                            # Query insert akhir
                            query = f"""
                                INSERT INTO slik_header ({columns})
                                VALUES ({placeholders})
                            """

                            cur.execute(query, all_values)

                        except Exception as e:
                            print(f"Error inserting data: {e}")
                            traceback.print_exc()
                            print(f"Item: {json_header}")
                            
                    uploaded_data = pd.DataFrame(json_header, index=[0])
                    table_data = uploaded_data.to_html(classes="table table-striped", index=False)
                    
                    if 'individual' in data:
                        
                        json_individual = data['individual']
                        json_paramPencarian = json_individual.get('parameterPencarian', {})
                        json_dpdebitur = json_individual.get('dataPokokDebitur', [])
                        json_rFasilitas = json_individual.get('ringkasanFasilitas', {})
                        fasilitas = json_individual.get('fasilitas', {})
                        json_fKreditPembiayan = fasilitas.get('kreditPembiayan', [])
                        json_fLC = fasilitas.get('lc', [])
                        json_fGaransi = fasilitas.get('garansiYgDiberikan', [])
                        json_fFasilitasLain = fasilitas.get('fasilitasLain', [])
                        nomor_laporan = json_individual.get('nomorLaporan')

                        # Hapus key yang ada saja
                        for key in ['dataPokokDebitur', 'parameterPencarian', 'ringkasanFasilitas', 'fasilitas']:
                            json_individual.pop(key, None)

                        uploaded_data_2 = pd.DataFrame(json_individual, index=[0]).fillna('')
                        uploaded_data_3 = pd.DataFrame(json_paramPencarian, index=[0]).fillna('')
                        uploaded_data_4 = pd.DataFrame(json_dpdebitur).fillna('')
                        uploaded_data_5 = pd.DataFrame(json_rFasilitas, index=[0]).fillna('')
                        
                        extra = {
                            'periodeData': json_posisiDataTerakhir,
                            'username': user_info['username'],
                            'namaFileUpload': user_info['nama_file'],
                            'uploadDate': current_datetime
                        }
                        
                        # Insert data ke database dengan pengecekan yang konsisten
                        if json_individual:
                            json_individual_clean = clean_data_for_sql(pd.DataFrame([json_individual])).to_dict('records')[0]
                            insert_data(cur, "slik_perusahaan", json_individual_clean, columns_to_remove, extra_columns=extra)

                        if json_paramPencarian:
                            json_paramPencarian_clean = clean_data_for_sql(pd.DataFrame([json_paramPencarian])).to_dict('records')[0]
                            insert_data(cur, "slik_parameter_pencarian", json_paramPencarian_clean, columns_to_remove, extra_columns=extra)

                        if json_dpdebitur:
                            for item in json_dpdebitur:
                                item_clean = clean_data_for_sql(pd.DataFrame([item])).to_dict('records')[0]
                                insert_data(cur, "slik_data_pokok_debitur", item_clean, columns_to_remove, extra_columns=extra)

                        if json_rFasilitas:
                            json_rFasilitas_clean = clean_data_for_sql(pd.DataFrame([json_rFasilitas])).to_dict('records')[0]
                            insert_data(cur, "slik_ringkasan_fasilitas", json_rFasilitas_clean, columns_to_remove, extra_columns=extra)
                        
                        if len(uploaded_files) > 1:
                        # table_data_6 = uploaded_data_6.to_html(classes="table table-striped", index=False)
                            all_uploaded_data.append(uploaded_data_2)
                            all_uploaded_data.append(uploaded_data_3)
                            all_uploaded_data.append(uploaded_data_4)
                            all_uploaded_data.append(uploaded_data_5)
                            # all_uploaded_data.append(uploaded_data_6)
                            
                            for data in all_uploaded_data:
                                table_html = data.to_html(classes="table table-striped", index=False).strip()
                                list_table_data.append(table_html)
                        
                        # Fasilitas Kredit Pembiayaan
                        if json_fKreditPembiayan:
                            for item in json_fKreditPembiayan:
                                item_clean = clean_data_for_sql(pd.DataFrame([item])).to_dict('records')[0]
                                insert_data(cur, "slik_fasilitas_kredit_pembiayaan", item_clean, columns_to_remove, extra_columns=extra)

                            uploaded_data_6 = pd.DataFrame(json_fKreditPembiayan)
                            uploaded_data_6.drop(columns=[col for col in columns_to_remove if col in uploaded_data_6.columns], inplace=True)
                            if len(uploaded_data_6) > 0:
                                uploaded_data_6 = uploaded_data_6.assign(**{'urutanFile': idx})
                            list_uploaded_data_6.append(uploaded_data_6)

                        # Fasilitas LC
                        if json_fLC:
                            for item in json_fLC:
                                item_clean = clean_data_for_sql(pd.DataFrame([item])).to_dict('records')[0]
                                insert_data(cur, "slik_fasilitas_lc", item_clean, columns_to_remove, extra_columns=extra)

                            uploaded_data_7 = pd.DataFrame(json_fLC)
                            uploaded_data_7.drop(columns=[col for col in columns_to_remove if col in uploaded_data_7.columns], inplace=True)
                            if len(uploaded_data_7) > 0:
                                uploaded_data_7 = uploaded_data_7.assign(**{'urutanFile': idx})
                            list_uploaded_data_7.append(uploaded_data_7)

                        # Fasilitas Garansi
                        if json_fGaransi:
                            for item in json_fGaransi:
                                item_clean = clean_data_for_sql(pd.DataFrame([item])).to_dict('records')[0]
                                insert_data(cur, "slik_fasilitas_garansi", item_clean, columns_to_remove, extra_columns=extra)

                            uploaded_data_8 = pd.DataFrame(json_fGaransi)
                            uploaded_data_8.drop(columns=[col for col in columns_to_remove if col in uploaded_data_8.columns], inplace=True)
                            if len(uploaded_data_8) > 0:
                                uploaded_data_8 = uploaded_data_8.assign(**{'urutanFile': idx})
                            list_uploaded_data_8.append(uploaded_data_8)

                        # Fasilitas Lainnya
                        if json_fFasilitasLain:
                            for item in json_fFasilitasLain:
                                item_clean = clean_data_for_sql(pd.DataFrame([item])).to_dict('records')[0]
                                insert_data(cur, "slik_fasilitas_lainnya", item_clean, columns_to_remove, extra_columns=extra)

                            uploaded_data_9 = pd.DataFrame(json_fFasilitasLain)
                            uploaded_data_9.drop(columns=[col for col in columns_to_remove if col in uploaded_data_9.columns], inplace=True)
                            if len(uploaded_data_9) > 0:
                                uploaded_data_9 = uploaded_data_9.assign(**{'urutanFile': idx})
                            list_uploaded_data_9.append(uploaded_data_9)
                            
                        # Set session data availability
                        if has_request_context():
                            session['data_available'] = True

                    elif 'perusahaan' in data:
                        
                        json_perusahaan = data['perusahaan']
                        json_paramPencarian = json_perusahaan.get('parameterPencarian', {})
                        json_dpdebitur = json_perusahaan.get('dataPokokDebitur', [])
                        json_rFasilitas = json_perusahaan.get('ringkasanFasilitas', {})
                        fasilitas = json_perusahaan.get('fasilitas', {})
                        
                        json_fKreditPembiayan = fasilitas.get('kreditPembiayan', [])
                        json_fSuratBerharga = fasilitas.get('suratBerharga', [])
                        json_fLC = fasilitas.get('lc', [])
                        json_fGaransi = fasilitas.get('garansiYgDiberikan', [])
                        json_fFasilitasLain = fasilitas.get('fasilitasLain', [])
                        json_kPengurusPemilik = json_perusahaan.get('kelompokPengurusPemilik', None)
                        nomor_laporan = json_perusahaan.get('nomorLaporan')

                        # Hapus key yang ada saja
                        for key in ['dataPokokDebitur', 'parameterPencarian', 'ringkasanFasilitas', 'fasilitas', 'kelompokPengurusPemilik']:
                            json_perusahaan.pop(key, None)
                        
                        uploaded_data_2 = pd.DataFrame(json_perusahaan, index=[0]).fillna('')
                        uploaded_data_3 = pd.DataFrame(json_paramPencarian, index=[0]).fillna('')
                        uploaded_data_4 = pd.DataFrame(json_dpdebitur).fillna('')
                        uploaded_data_5 = pd.DataFrame(json_rFasilitas, index=[0]).fillna('')
                        
                        extra = {
                            'periodeData': json_posisiDataTerakhir,
                            'username': user_info['username'],
                            'namaFileUpload': user_info['nama_file'],
                            'uploadDate': current_datetime
                        }
                        
                        if json_perusahaan:
                            json_perusahaan_clean = clean_data_for_sql(pd.DataFrame([json_perusahaan])).to_dict('records')[0]
                            insert_data(cur, "slik_perusahaan", {
                                "nomorLaporan": json_perusahaan_clean.get("nomorLaporan"),
                                "posisiDataTerakhir": json_perusahaan_clean.get("posisiDataTerakhir"),
                                "tanggalPermintaan": json_perusahaan_clean.get("tanggalPermintaan")
                            }, extra_columns=extra)

                        if json_paramPencarian:
                            json_paramPencarian_clean = clean_data_for_sql(pd.DataFrame([json_paramPencarian])).to_dict('records')[0]
                            insert_data(cur, "slik_parameter_pencarian", json_paramPencarian_clean, columns_to_remove, extra_columns=extra)

                        if json_dpdebitur:
                            for item in json_dpdebitur:
                                item_clean = clean_data_for_sql(pd.DataFrame([item])).to_dict('records')[0]
                                insert_data(cur, "slik_data_pokok_debitur", item_clean, columns_to_remove, extra_columns=extra)

                        if json_kPengurusPemilik:
                            for kelompok in json_kPengurusPemilik:
                                kelompok_clean = clean_data_for_sql(pd.DataFrame([kelompok])).to_dict('records')[0]
                                for pengurus in kelompok["pengurusPemilik"]:
                                    pengurus_clean = clean_data_for_sql(pd.DataFrame([pengurus])).to_dict('records')[0]
                                    full_item = {
                                        "kodeLJK": kelompok_clean["kodeLJK"],
                                        "namaLJK": kelompok_clean["namaLJK"],
                                        **pengurus_clean
                                    }
                                    insert_data(cur, "slik_kelompok_pengurus_pemilik", full_item, columns_to_remove, extra_columns=extra)

                        if json_rFasilitas:
                            json_rFasilitas_clean = clean_data_for_sql(pd.DataFrame([json_rFasilitas])).to_dict('records')[0]
                            insert_data(cur, "slik_ringkasan_fasilitas", json_rFasilitas_clean, columns_to_remove, extra_columns=extra)
                            
                        if len(uploaded_files) > 1:
                            all_uploaded_data.append(uploaded_data_2)
                            all_uploaded_data.append(uploaded_data_3)
                            all_uploaded_data.append(uploaded_data_4)
                            all_uploaded_data.append(uploaded_data_5)
                            
                            for data in all_uploaded_data:
                                table_html = data.to_html(classes="table table-striped", index=False).strip()
                                list_table_data.append(table_html)
                        
                        if json_fKreditPembiayan:
                            for item in json_fKreditPembiayan:
                                item_clean = clean_data_for_sql(pd.DataFrame([item])).to_dict('records')[0]
                                insert_data(cur, "slik_fasilitas_kredit_pembiayaan", item_clean, columns_to_remove, extra_columns=extra)

                            uploaded_data_6 = pd.DataFrame(json_fKreditPembiayan)
                            uploaded_data_6.drop(columns=[col for col in columns_to_remove if col in uploaded_data_6.columns], inplace=True)
                            if len(uploaded_data_6) > 0:
                                uploaded_data_6 = uploaded_data_6.assign(**{'urutanFile': idx})
                            list_uploaded_data_6.append(uploaded_data_6)

                        # Fasilitas LC
                        if json_fLC:
                            for item in json_fLC:
                                item_clean = clean_data_for_sql(pd.DataFrame([item])).to_dict('records')[0]
                                insert_data(cur, "slik_fasilitas_lc", item_clean, columns_to_remove, extra_columns=extra)

                            uploaded_data_7 = pd.DataFrame(json_fLC)
                            uploaded_data_7.drop(columns=[col for col in columns_to_remove if col in uploaded_data_7.columns], inplace=True)
                            if len(uploaded_data_7) > 0:
                                uploaded_data_7 = uploaded_data_7.assign(**{'urutanFile': idx})
                            list_uploaded_data_7.append(uploaded_data_7)

                        # Fasilitas Garansi
                        if json_fGaransi:
                            for item in json_fGaransi:
                                item_clean = clean_data_for_sql(pd.DataFrame([item])).to_dict('records')[0]
                                insert_data(cur, "slik_fasilitas_garansi", item_clean, columns_to_remove, extra_columns=extra)

                            uploaded_data_8 = pd.DataFrame(json_fGaransi)
                            uploaded_data_8.drop(columns=[col for col in columns_to_remove if col in uploaded_data_8.columns], inplace=True)
                            if len(uploaded_data_8) > 0:
                                uploaded_data_8 = uploaded_data_8.assign(**{'urutanFile': idx})
                            list_uploaded_data_8.append(uploaded_data_8)

                        # Fasilitas Lainnya
                        if json_fFasilitasLain:
                            for item in json_fFasilitasLain:
                                item_clean = clean_data_for_sql(pd.DataFrame([item])).to_dict('records')[0]
                                insert_data(cur, "slik_fasilitas_lainnya", item_clean, columns_to_remove, extra_columns=extra)

                            uploaded_data_9 = pd.DataFrame(json_fFasilitasLain)
                            uploaded_data_9.drop(columns=[col for col in columns_to_remove if col in uploaded_data_9.columns], inplace=True)
                            if len(uploaded_data_9) > 0:
                                uploaded_data_9 = uploaded_data_9.assign(**{'urutanFile': idx})
                            list_uploaded_data_9.append(uploaded_data_9)

                        # Surat Berharga
                        if json_fSuratBerharga:
                            for item in json_fSuratBerharga:
                                item_clean = clean_data_for_sql(pd.DataFrame([item])).to_dict('records')[0]
                                insert_data(cur, "slik_fasilitas_surat_berharga", item_clean, columns_to_remove, extra_columns=extra)

                            uploaded_data_10 = pd.DataFrame(json_fSuratBerharga)
                            uploaded_data_10.drop(columns=[col for col in columns_to_remove if col in uploaded_data_10.columns], inplace=True)
                            if len(uploaded_data_10) > 0:
                                uploaded_data_10 = uploaded_data_10.assign(**{'urutanFile': idx})
                            list_uploaded_data_10.append(uploaded_data_10)      
                                            
                        if json_kPengurusPemilik:
                            df_kPengurusPemilik = pd.DataFrame(json_kPengurusPemilik) 
                            data_temp = {'kodeLJK': ['1', '2', '3'], 'namaLJK': ['A', 'B', 'C'], 'pengurusPemilik': ['X', 'Y', 'Z']}
                            df_temp = pd.DataFrame(data_temp)
                            df_expanded = df_temp.head(0)

                            for row in df_kPengurusPemilik.itertuples(index=False):
                                for x in row.pengurusPemilik:
                                    df_expanded.loc[len(df_expanded)] = [row.kodeLJK, row.namaLJK, x]
                        
                            df_expanded = df_expanded.join(pd.json_normalize(df_expanded.pop('pengurusPemilik')))

                            uploaded_data_11 = df_expanded
                    if has_request_context():
                        session['data_available'] = True
                    
                except Exception as e:
                    app.logger.error(f"[{task_id}] Error processing file {filename}: {str(e)}")
                    return {
                        "error": True,
                        "error_type": "system_error",
                        "message": f"Error processing file {filename}: {str(e)}",
                        "redirect_url": "/upload-big-size"
                    }
            else:
                # task_progress[task_id]['status'] = 'error'
                return {
                    "error": True,
                    "error_type": "validation_error",
                    "message": "Please upload a valid Text file",
                    "redirect_url": "/upload-big-size"
                }
        
        task_progress[task_id]['progress'] = 90
        
        if len(list_uploaded_data_6) > 0:
            missing_ljk = [i for i, df in enumerate(list_uploaded_data_6) if 'ljk' not in df.columns]
            if not missing_ljk:
                uploaded_data_4_dedup = uploaded_data_4.drop_duplicates(subset=['pelapor', 'pelaporKet'])
                combined_data_6 = pd.concat(list_uploaded_data_6, ignore_index=True)
                merged_fKP = combined_data_6.merge(
                    uploaded_data_4,
                    left_on=['ljk', 'ljkKet'],
                    right_on=['pelapor', 'pelaporKet'],
                    how='left'
                )
                
                if 'npwp' in merged_fKP.columns and 'noIdentitas' in merged_fKP.columns:
                    if 'individual' in data:
                        # Untuk individual: prioritas noIdentitas, fallback ke npwp
                        merged_fKP['npwp'] = merged_fKP.apply(
                            lambda row: row['noIdentitas'] if pd.notna(row['noIdentitas']) and str(row['noIdentitas']).strip() != '' 
                                    else (row['npwp'] if pd.notna(row['npwp']) and str(row['npwp']).strip() != '' else None),
                            axis=1
                        )
                    elif 'perusahaan' in data:
                        # Untuk perusahaan: prioritas npwp, fallback ke noIdentitas
                        merged_fKP['npwp'] = merged_fKP.apply(
                            lambda row: row['npwp'] if pd.notna(row['npwp']) and str(row['npwp']).strip() != '' 
                                    else (row['noIdentitas'] if pd.notna(row['noIdentitas']) and str(row['noIdentitas']).strip() != '' else None),
                            axis=1
                        )
                elif 'noIdentitas' in merged_fKP.columns and 'npwp' not in merged_fKP.columns:
                    # Jika hanya ada noIdentitas, rename ke npwp
                    merged_fKP = merged_fKP.rename(columns={'noIdentitas': 'npwp'})
                
                # ACTIVE FACILITY (Kondisi == '00')
                active_fKP = merged_fKP[merged_fKP['kondisi'].isin(['00', '03', '13', '16'])]
            
                if not active_fKP.empty:
                    # Daftar kolom yang dibutuhkan
                    columns = [
                        'namaDebitur', 'npwp', 'alamat', 'ljkKet',
                        'jenisKreditPembiayaanKet', 'jenisPenggunaanKet', 'plafon',
                        'bakiDebet', 'tunggakanPokok', 'tunggakanBunga', 'denda',
                        'jumlahHariTunggakan', 'kualitas', 'kualitasKet', 'tahunBulan24', 'urutanFile'
                    ]

                    # Tambahkan kolom opsional jika tersedia
                    if 'tglAktaPendirian' in active_fKP.columns:
                        columns.insert(2, 'tglAktaPendirian')
                    elif 'tanggalLahir' in active_fKP.columns:
                        base_columns.insert(2, 'tanggalLahir')
                        
                    if 'valutaKode' in active_fKP.columns:
                        columns.insert(12, 'valutaKode')

                    # Filter kolom yang ada
                    available_columns = [col for col in columns if col in active_fKP.columns]
                    active_facility_1 = active_fKP[available_columns].copy()
                    active_facility_1 = clean_data_for_sql(active_facility_1)
                    
                    # Tambahkan kolom metadata
                    active_facility_1['periodeData'] = json_posisiDataTerakhir
                    active_facility_1['username'] = username
                    active_facility_1['namaFileUpload'] = nama_file
                    active_facility_1['uploadDate'] = current_datetime
                    
                    # Rename dict untuk tampilan HTML
                    rename_dict = {
                        'namaDebitur': 'Nama Debitur/Calon Debitur',
                        'npwp': 'Nomor Identitas',
                        'alamat': 'Alamat',
                        'ljkKet': 'Kreditur/Pelapor',
                        'jenisKreditPembiayaanKet': 'Jenis Kredit/Pembiayaan',
                        'jenisPenggunaanKet': 'Jenis Penggunaan',
                        'plafon': 'Plafon',
                        'bakiDebet': 'Oustanding/Baki Debet',
                        'tunggakanPokok': 'Tunggakan Pokok',
                        'tunggakanBunga': 'Tunggakan Bunga',
                        'denda': 'Denda',
                        'jumlahHariTunggakan': 'Hari Keterlambatan',
                        'kualitas': 'Kode Kolektibilitas Saat ini',
                        'kualitasKet': 'Kolektibilitas Saat ini',
                        'tahunBulan24': 'Periode Pelaporan Terakhir',
                        'urutanFile': 'File ke'
                    }
                    if 'tglAktaPendirian' in active_fKP.columns:
                        rename_dict['tglAktaPendirian'] = 'Tanggal Pendirian'
                    elif 'tanggalLahir' in active_fKP.columns:
                        rename_dict['tanggalLahir'] = 'Tanggal Lahir'

                    
                    if 'valutaKode' in active_fKP.columns:
                        rename_dict['valutaKode'] = 'Valuta'
                    
                    # === Persiapkan data summary ===
                    summary_columns = [
                        'periodeData', 'username', 'namaFileUpload', 'uploadDate',
                        'namaDebitur', 'nomorLaporan', 'nomorIdentitas', 'kreditur',
                        'jenisPembiayaan', 'kodeKolektibilitas', 'kolektibilitas'
                    ]

                    # Buat dictionary untuk pemetaan kolom
                    summary_map = {
                        'periodeData': 'periodeData', 
                        'username': 'username',
                        'namaFileUpload': 'namaFileUpload', 
                        'uploadDate': 'uploadDate',
                        'namaDebitur': 'namaDebitur', 
                        'nomorLaporan': None,
                        'nomorIdentitas': 'npwp', 
                        'kreditur': 'ljkKet',
                        'jenisPembiayaan': 'jenisKreditPembiayaanKet', 
                        'kodeKolektibilitas': 'kualitas',
                        'kolektibilitas': 'kualitasKet'
                    }
                    
                    # Buat DataFrame summary baru
                    summary_data = pd.DataFrame()
                    for target_col, source_col in summary_map.items():
                        if source_col is None:
                            if target_col == 'nomorLaporan':
                                summary_data[target_col] = nomor_laporan
                        elif source_col in active_facility_1.columns:
                            summary_data[target_col] = active_facility_1[source_col]
                        else:
                            summary_data[target_col] = None
                    
                    # Konversi NaN ke None untuk semua DataFrame
                    summary_data = clean_data_for_sql(summary_data)
                    
                    # Insert data summary
                    try:
                        summary_columns_sql = ', '.join(summary_columns)
                        summary_placeholders = ', '.join(['?'] * len(summary_columns))
                        summary_query = f"""
                            INSERT INTO slik_summary_fasilitas_aktif ({summary_columns_sql})
                            VALUES ({summary_placeholders})
                        """
                        
                        summary_values = list(summary_data.itertuples(index=False, name=None))
                        if summary_values:
                            print(f"Inserting {len(summary_values)} records into slik_summary_fasilitas_aktif")
                            cur.executemany(summary_query, summary_values)
                            print("Data summary berhasil di-insert!")
                        else:
                            print("Tidak ada data summary untuk di-insert!")
                    except Exception as e:
                        print(f"Error saat insert data summary: {str(e)}")
                    
                    # Insert ke tabel fasilitas aktif
                    try:
                        columns_af = ', '.join(active_facility_1.columns)
                        placeholders = ', '.join(['?'] * len(active_facility_1.columns))
                        query = f"""
                            INSERT INTO slik_fasilitas_aktif_kredit_pembiayaan ({columns_af})
                            VALUES ({placeholders})
                        """
                        
                        data_insert = list(active_facility_1.itertuples(index=False, name=None))  # ← GANTI NAMA
                        if data_insert:  # ← GANTI NAMA
                            cur.executemany(query, data_insert)  # ← GANTI NAMA
                            print("Data aktif berhasil di-insert!")
                        else:
                            print("Data fasilitas aktif kosong!")
                    except Exception as e:
                        print(f"Error saat insert data aktif: {str(e)}")

                    # Persiapkan data untuk tabel HTML
                    active_facility_display = active_facility_1.copy()
                    active_facility_display = active_facility_display.rename(columns=rename_dict)
                    active_facility_display.insert(1, 'Nomor Laporan', nomor_laporan)
                    active_facility_display = active_facility_display.reset_index(names='No')
                    active_facility_display['No'] = active_facility_display.index + 1
                    active_facility_display = active_facility_display.drop(columns=['periodeData', 'username', 'namaFileUpload', 'uploadDate'], errors='ignore')

                # CLOSED FACILITY (Kondisi == '02')
                closed_fKP = merged_fKP[merged_fKP['kondisi'].isin(['01', '02', '04', '05', '06', '07', '08', '09', '12', '17'])]
                columns_closed = [
                    'namaDebitur', 'npwp', 'alamat', 'ljkKet',
                    'jenisKreditPembiayaanKet', 'jenisPenggunaanKet', 'plafon',
                    'bakiDebet', 'tunggakanPokok', 'tunggakanBunga', 'denda',
                    'jumlahHariTunggakan', 'kualitas', 'kualitasKet', 'tahunBulan24', 'urutanFile'
                ]
                
                if 'npwp' not in closed_fKP.columns and 'noIdentitas' in closed_fKP.columns:
                    closed_fKP = closed_fKP.rename(columns={'noIdentitas': 'npwp'})
                elif 'identitas' in closed_fKP.columns and 'npwp' not in closed_fKP.columns:
                    closed_fKP = closed_fKP.rename(columns={'identitas': 'npwp'})

                if 'tglAktaPendirian' in closed_fKP.columns:
                    columns_closed.insert(2, 'tglAktaPendirian')
                elif 'tanggalLahir' in closed_fKP.columns:    
                    columns_closed.insert(2, 'tanggalLahir')
                    
                if 'valutaKode' in closed_fKP.columns:
                    columns_closed.insert(12, 'valutaKode')  # Fixed here - was using 'columns' instead of 'columns_closed'

                available_columns_closed = [col for col in columns_closed if col in closed_fKP.columns]
                closed_facility_1 = closed_fKP[available_columns_closed].copy()
                closed_facility_1 = clean_data_for_sql(closed_facility_1)
                rename_dict_closed = {
                    'namaDebitur': 'Nama Debitur/Calon Debitur',
                    'npwp': 'Nomor Identitas',
                    'alamat': 'Alamat',
                    'ljkKet': 'Kreditur/Pelapor',
                    'jenisKreditPembiayaanKet': 'Jenis Kredit/Pembiayaan',
                    'jenisPenggunaanKet': 'Jenis Penggunaan',
                    'plafon': 'Plafon',
                    'bakiDebet': 'Oustanding/Baki Debet',
                    'tunggakanPokok': 'Tunggakan Pokok',
                    'tunggakanBunga': 'Tunggakan Bunga',
                    'denda': 'Denda',
                    'jumlahHariTunggakan': 'Hari Keterlambatan',
                    'kualitas': 'Kode Kolektibilitas Saat ini',
                    'kualitasKet': 'Kolektibilitas Saat ini',
                    'tahunBulan24': 'Periode Pelaporan Terakhir',
                    'urutanFile': 'File ke'
                }

                if 'tglAktaPendirian' in closed_fKP.columns:
                    rename_dict_closed['tglAktaPendirian'] = 'Tanggal Pendirian'
                elif 'tanggalLahir' in closed_fKP.columns:
                    rename_dict_closed['tanggalLahir'] = 'Tanggal Lahir'
                    
                if 'valutaKode' in closed_fKP.columns:
                    rename_dict_closed['valutaKode'] = 'Valuta'  # Fixed here - was using 'rename_dict' instead of 'rename_dict_closed'
                    
                if 'tanggalLahir' in closed_facility_1.columns:
                    closed_facility_1 = closed_facility_1.rename(columns={'tanggalLahir': 'tglAktaPendirian'})
                
                closed_facility_1['periodeData'] = json_posisiDataTerakhir
                closed_facility_1['username'] = username
                closed_facility_1['namaFileUpload'] = nama_file
                closed_facility_1['uploadDate'] = current_datetime
                
                # Ambil nama kolom dari DataFrame
                columns_cf = ', '.join(closed_facility_1.columns)  # Changed variable name to avoid conflict

                # Placeholder SQL Server pakai '?'
                placeholders = ', '.join(['?'] * len(closed_facility_1.columns))

                # Buat query insert
                query = f"""
                    INSERT INTO slik_fasilitas_lunas_kredit_pembiayaan ({columns_cf})
                    VALUES ({placeholders})
                """
                # Ubah DataFrame ke list of tuple tanpa index
                data_closed = list(closed_facility_1.itertuples(index=False, name=None))  # ← GANTI NAMA

                # Jalankan batch insert
                if data_closed:  # ← GANTI NAMA
                    cur.executemany(query, data_closed)  # ← GANTI NAMA
                else:
                    print("Data kosong!")
                
                closed_facility_1 = closed_facility_1.rename(columns=rename_dict_closed)
                closed_facility_1.insert(1, 'Nomor Laporan', nomor_laporan, allow_duplicates=False)  # Added this line to match active_facility_1 treatment
                closed_facility_1 = closed_facility_1.reset_index(names='No')
                closed_facility_1['No'] = closed_facility_1.index + 1
                closed_facility_1 = closed_facility_1.drop(columns=['periodeData', 'username', 'namaFileUpload', 'uploadDate'], errors='ignore')
        
        if len(list_uploaded_data_7) > 0:
            # Periksa apakah semua DataFrame dalam list memiliki kolom 'ljk'
            missing_ljk = [i for i, df in enumerate(list_uploaded_data_7) if 'ljk' not in df.columns]
            if not missing_ljk:
                # Proses data umum
                def process_data(df_filtered, table_name, summary_table=None):
                    # Kolom dasar yang dibutuhkan
                    base_columns = [
                        'namaDebitur', 'npwp', 'alamat',
                        'ljkKet', 'jenisLcKet', 'tujuanLcKet', 'plafon', 'nominalLc',
                        'tanggalWanPrestasi', 'kualitas', 'kualitasKet',
                        'tahunBulan24', 'urutanFile'
                    ]
                    
                    if 'tanggalLahir' in df_filtered.columns:
                        df_filtered = df_filtered.rename(columns={'tanggalLahir': 'tglAktaPendirian'})
                    
                    if 'tglAktaPendirian' in df_filtered.columns:
                        base_columns.insert(2, 'tglAktaPendirian')
                    
                    # Tambahkan kolom 'valuta' jika tersedia
                    if 'valuta' in df_filtered.columns:
                        base_columns.insert(9, 'valuta')

                    # Siapkan DataFrame untuk disimpan ke database
                    available_cols = [col for col in base_columns if col in df_filtered.columns]
                    result_df = df_filtered[available_cols].copy()
                    result_df = clean_data_for_sql(result_df)
                
                    result_df['periodeData'] = json_posisiDataTerakhir
                    result_df['username'] = username
                    result_df['namaFileUpload'] = nama_file
                    result_df['uploadDate'] = current_datetime
                    
                    # Insert ke database
                    if not result_df.empty:
                        columns = ', '.join(result_df.columns)
                        placeholders = ', '.join(['?'] * len(result_df.columns))
                        query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
                        data_insert = list(result_df.itertuples(index=False, name=None))
                        cur.executemany(query, data_insert)
                    
                    # Tambahkan ke tabel summary jika diperlukan
                    if summary_table and not result_df.empty:
                        summary_data = result_df.copy()
                        summary_data = clean_data_for_sql(summary_data)
                        summary_data['nomorLaporan'] = nomor_laporan
                        summary_data.rename(columns={
                            'namaDebitur': 'namaDebitur',
                            'npwp': 'nomorIdentitas',
                            'ljkKet': 'kreditur',
                            'jenisLcKet': 'jenisPembiayaan',
                            'kualitas': 'kodeKolektibilitas',
                            'kualitasKet': 'kolektibilitas'
                        }, inplace=True)
                        
                        summary_columns = [
                            'periodeData', 'username', 'namaFileUpload', 'uploadDate',
                            'namaDebitur', 'nomorLaporan', 'nomorIdentitas',
                            'kreditur', 'jenisPembiayaan', 'kodeKolektibilitas', 'kolektibilitas'
                        ]
                        
                        summary_df = summary_data[summary_columns]
                        summary_query = f"""
                            INSERT INTO {summary_table} ({', '.join(summary_columns)})
                            VALUES ({', '.join(['?'] * len(summary_columns))})
                        """
                        summary_tuples = list(summary_df.itertuples(index=False, name=None))
                        cur.executemany(summary_query, summary_tuples)
                    
                    # Format data untuk HTML display
                    column_rename_map = {
                        'namaDebitur': 'Nama Debitur/Calon Debitur',
                        'npwp': 'Nomor Identitas',
                        'tglAktaPendirian': 'Tanggal Lahir/Pendirian',
                        'alamat': 'Alamat',
                        'ljkKet': 'Kreditur/Pelapor',
                        'jenisLcKet': 'Jenis L/C',
                        'tujuanLcKet': 'Tujuan L/C',
                        'plafon': 'Plafon',
                        'nominalLc': 'Oustanding/Baki Debet',
                        'valuta': 'Valuta',
                        'tanggalWanPrestasi': 'Tanggal Wan prestasi',
                        'kualitas': 'Kode Kolektibilitas Saat ini',
                        'kualitasKet': 'Kolektibilitas Saat ini',
                        'tahunBulan24': 'Periode Pelaporan Terakhir',
                        'urutanFile': 'File ke'
                    }
                    
                    display_df = result_df.rename(columns=column_rename_map)
                    if summary_table:  # Hanya tambahkan nomor laporan jika ini adalah fasilitas aktif
                        display_df.insert(1, 'Nomor Laporan', nomor_laporan)
                    display_df.reset_index(drop=True, inplace=True)
                    display_df.insert(0, 'No', range(1, len(display_df) + 1))
                    display_df = display_df.drop(columns=['periodeData', 'username', 'namaFileUpload', 'uploadDate'], errors='ignore')
                    
                    return display_df.to_html(classes="table table-striped", index=False)
                
                # Deduplicate dan gabungkan data
                uploaded_data_4_dedup = uploaded_data_4.drop_duplicates(subset=['pelapor', 'pelaporKet'])
                combined_data_7 = pd.concat(list_uploaded_data_7, ignore_index=True)
                merged_fLC = combined_data_7.merge(
                    uploaded_data_4_dedup,
                    left_on=['ljk', 'ljkKet'],
                    right_on=['pelapor', 'pelaporKet'],
                    how='left'
                )
                
                if 'npwp' in merged_fLC.columns and 'noIdentitas' in merged_fLC.columns:
                    if 'individual' in data:
                        # Untuk individual: prioritas noIdentitas, fallback ke npwp
                        merged_fLC['npwp'] = merged_fLC.apply(
                            lambda row: row['noIdentitas'] if pd.notna(row['noIdentitas']) and str(row['noIdentitas']).strip() != '' 
                                    else (row['npwp'] if pd.notna(row['npwp']) and str(row['npwp']).strip() != '' else None),
                            axis=1
                        )
                    elif 'perusahaan' in data:
                        # Untuk perusahaan: prioritas npwp, fallback ke noIdentitas
                        merged_fLC['npwp'] = merged_fLC.apply(
                            lambda row: row['npwp'] if pd.notna(row['npwp']) and str(row['npwp']).strip() != '' 
                                    else (row['noIdentitas'] if pd.notna(row['noIdentitas']) and str(row['noIdentitas']).strip() != '' else None),
                            axis=1
                        )
                elif 'noIdentitas' in merged_fLC.columns and 'npwp' not in merged_fLC.columns:
                    # Jika hanya ada noIdentitas, rename ke npwp
                    merged_fLC = merged_fLC.rename(columns={'noIdentitas': 'npwp'})
                
                # Proses fasilitas aktif
                active_fLC = merged_fLC[merged_fLC['kondisi'].isin(['00', '03', '13', '16'])]
                process_data(
                    active_fLC, 
                    'slik_fasilitas_aktif_lc', 
                    'slik_summary_fasilitas_aktif'
                )
                
                # Proses fasilitas lunas
                closed_fLC = merged_fLC[merged_fLC['kondisi'].isin(['01', '02', '04', '05', '06', '07', '08', '09', '12', '17'])]
                process_data(
                    closed_fLC, 
                    'slik_fasilitas_lunas_lc'
                )
                    
        if len(list_uploaded_data_8) > 0:
            missing_ljk = [i for i, df in enumerate(list_uploaded_data_8) if 'ljk' not in df.columns]
            if not missing_ljk:
                # Fungsi umum untuk memproses dan menyimpan data
                def process_guarantee_data(df_filtered, table_name, is_active=False):
                    # Kolom dasar yang dibutuhkan
                    base_columns = [
                        'namaDebitur', 'npwp', 'alamat',
                        'ljkKet', 'jenisGaransiKet', 'tujuanGaransiKet', 'plafon',
                        'nominalBg', 'tanggalWanPrestasi', 'kualitas',
                        'kualitasKet', 'tahunBulan24', 'urutanFile'
                    ]
                    
                    if 'tanggalLahir' in df_filtered.columns:
                        df_filtered = df_filtered.rename(columns={'tanggalLahir': 'tglAktaPendirian'})
                    elif 'tglAktaPendirian' in df_filtered.columns:
                        base_columns.insert(2, 'tglAktaPendirian')
                    
                    # Tambahkan kolom valuta jika tersedia (dengan penanganan nama kolom yang berbeda)
                    valuta_column = 'kodeValuta' if 'kodeValuta' in df_filtered.columns else 'valutaKode'
                    if valuta_column in df_filtered.columns:
                        base_columns.insert(9, valuta_column)
                    
                    # Siapkan DataFrame untuk disimpan ke database
                    available_cols = [c for c in base_columns if c in df_filtered.columns]
                    result_df = df_filtered[available_cols].copy()
                    result_df = clean_data_for_sql(result_df)
                    
                    result_df['periodeData'] = json_posisiDataTerakhir
                    result_df['username'] = username
                    result_df['namaFileUpload'] = nama_file
                    result_df['uploadDate'] = current_datetime
                    
                    # Insert ke database
                    if not result_df.empty:
                        columns = ', '.join(result_df.columns)
                        placeholders = ', '.join(['?'] * len(result_df.columns))
                        query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
                        data_insert = list(result_df.itertuples(index=False, name=None))
                        cur.executemany(query, data_insert)
                    else:
                        print("Data kosong!")
                    
                    # Jika aktif, tambahkan ke tabel summary
                    if is_active and not result_df.empty:
                        summary_data = result_df[[
                            'namaDebitur', 'npwp', 'ljkKet', 'jenisGaransiKet', 'kualitas', 'kualitasKet'
                        ]].copy()
                        summary_data = clean_data_for_sql(summary_data)
                        
                        summary_data['periodeData'] = json_posisiDataTerakhir
                        summary_data['username'] = username
                        summary_data['namaFileUpload'] = nama_file
                        summary_data['uploadDate'] = current_datetime
                        summary_data['nomorLaporan'] = nomor_laporan
                        
                        summary_data = summary_data.rename(columns={
                            'namaDebitur': 'namaDebitur',
                            'npwp': 'nomorIdentitas',
                            'ljkKet': 'kreditur',
                            'jenisGaransiKet': 'jenisPembiayaan',
                            'kualitas': 'kodeKolektibilitas',
                            'kualitasKet': 'kolektibilitas'
                        })
                        
                        summary_columns = [
                            'periodeData', 'username', 'namaFileUpload', 'uploadDate',
                            'namaDebitur', 'nomorLaporan', 'nomorIdentitas',
                            'kreditur', 'jenisPembiayaan', 'kodeKolektibilitas', 'kolektibilitas'
                        ]
                        
                        summary_df = summary_data[summary_columns]
                        summary_query = f"""
                            INSERT INTO slik_summary_fasilitas_aktif ({', '.join(summary_columns)})
                            VALUES ({', '.join(['?'] * len(summary_columns))})
                        """
                        summary_tuples = list(summary_df.itertuples(index=False, name=None))
                        
                        if summary_tuples:
                            cur.executemany(summary_query, summary_tuples)
                    
                    # Format data untuk HTML display
                    column_rename_map = {
                        'namaDebitur': 'Nama Debitur/Calon Debitur',
                        'npwp': 'Nomor Identitas',
                        'tglAktaPendirian': 'Tanggal Lahir/Pendirian',
                        'alamat': 'Alamat',
                        'ljkKet': 'Kreditur/Pelapor',
                        'jenisGaransiKet': 'Jenis Garansi',
                        'tujuanGaransiKet': 'Tujuan Garansi',
                        'plafon': 'Plafon',
                        'nominalBg': 'Oustanding/Baki Debet',
                        'kodeValuta': 'Valuta',
                        'valutaKode': 'Valuta',
                        'tanggalWanPrestasi': 'Tanggal Wan Prestasi',
                        'kualitas': 'Kode Kolektibilitas Saat ini',
                        'kualitasKet': 'Kolektibilitas Saat ini',
                        'tahunBulan24': 'Periode Pelaporan Terakhir',
                        'urutanFile': 'File ke'
                    }
                    
                    display_df = result_df.rename(columns=column_rename_map)
                    display_df.reset_index(drop=True, inplace=True)
                    
                    # Tambahkan Nomor Laporan untuk fasilitas aktif
                    if is_active:
                        display_df.insert(1, 'Nomor Laporan', nomor_laporan)
                    
                    # Tambahkan nomor urut
                    display_df.insert(0, 'No', range(1, len(display_df) + 1))
                    
                    # Hapus kolom internal
                    display_df = display_df.drop(columns=['periodeData', 'username', 'namaFileUpload', 'uploadDate'], errors='ignore')
                    
                    return display_df.to_html(classes="table table-striped", index=False)
                
                # Persiapan data
                uploaded_data_4_dedup = uploaded_data_4.drop_duplicates(subset=['pelapor', 'pelaporKet'])
                combined_data_8 = pd.concat(list_uploaded_data_8, ignore_index=True)
                merged_fGar = combined_data_8.merge(
                    uploaded_data_4_dedup,
                    left_on=['ljk', 'ljkKet'],
                    right_on=['pelapor', 'pelaporKet'],
                    how='left'
                )
                
                if 'npwp' in merged_fGar.columns and 'noIdentitas' in merged_fGar.columns:
                    if 'individual' in data:
                        # Untuk individual: prioritas noIdentitas, fallback ke npwp
                        merged_fGar['npwp'] = merged_fGar.apply(
                            lambda row: row['noIdentitas'] if pd.notna(row['noIdentitas']) and str(row['noIdentitas']).strip() != '' 
                                    else (row['npwp'] if pd.notna(row['npwp']) and str(row['npwp']).strip() != '' else None),
                            axis=1
                        )
                    elif 'perusahaan' in data:
                        # Untuk perusahaan: prioritas npwp, fallback ke noIdentitas
                        merged_fGar['npwp'] = merged_fGar.apply(
                            lambda row: row['npwp'] if pd.notna(row['npwp']) and str(row['npwp']).strip() != '' 
                                    else (row['noIdentitas'] if pd.notna(row['noIdentitas']) and str(row['noIdentitas']).strip() != '' else None),
                            axis=1
                        )
                elif 'noIdentitas' in merged_fGar.columns and 'npwp' not in merged_fGar.columns:
                    # Jika hanya ada noIdentitas, rename ke npwp
                    merged_fGar = merged_fGar.rename(columns={'noIdentitas': 'npwp'})
                
                # Proses fasilitas aktif
                active_fGar = merged_fGar[merged_fGar['kodeKondisi'].isin(['00', '03', '13', '16'])]
                process_guarantee_data(
                    active_fGar, 
                    'slik_fasilitas_aktif_bank_garansi',
                    is_active=True
                )
                
                # Proses fasilitas lunas
                closed_fGar = merged_fGar[merged_fGar['kodeKondisi'].isin(['01', '02', '04', '05', '06', '07', '08', '09', '12', '17'])]
                process_guarantee_data(
                    closed_fGar, 
                    'slik_fasilitas_lunas_bank_garansi'
                )
                    
        if len(list_uploaded_data_9) > 0:
            missing_ljk = [i for i, df in enumerate(list_uploaded_data_9) if 'ljk' not in df.columns]
            if not missing_ljk:
                # Fungsi umum untuk memproses dan menyimpan data fasilitas lainnya
                def process_other_facility(df_filtered, table_name, is_active=False):
                    # Kolom dasar yang dibutuhkan
                    base_columns = [
                        'namaDebitur', 'npwp', 'alamat',
                        'ljkKet', 'jenisFasilitasKet', 'nominalJumlahKwajibanIDR',
                        'jumlahHariTunggakan', 'kualitas',
                        'kualitasKet', 'tahunBulan24', 'urutanFile'
                    ]
                    
                    if 'tanggalLahir' in df_filtered.columns:
                        df_filtered = df_filtered.rename(columns={'tanggalLahir': 'tglAktaPendirian'})
                    if 'tglAktaPendirian' in df_filtered.columns:
                        base_columns.insert(2, 'tglAktaPendirian')
                    
                    # Tambahkan kolom valuta jika tersedia
                    if 'kodeValuta' in df_filtered.columns:
                        base_columns.insert(7, 'kodeValuta')
                    
                    # Siapkan DataFrame untuk disimpan ke database
                    available_cols = [c for c in base_columns if c in df_filtered.columns]
                    result_df = df_filtered[available_cols].copy()
                    result_df = clean_data_for_sql(result_df)
                    
                    # Tambahkan kolom metadata
                    metadata_columns = {
                        'periodeData': json_posisiDataTerakhir,
                        'username': username,
                        'namaFileUpload': nama_file,
                        'uploadDate': current_datetime
                    }
                    
                    for col, val in metadata_columns.items():
                        result_df[col] = val
                    
                    # Insert ke database utama jika ada data
                    if not result_df.empty:
                        columns = ', '.join(result_df.columns)
                        placeholders = ', '.join(['?'] * len(result_df.columns))
                        query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
                        data_insert = list(result_df.itertuples(index=False, name=None))
                        cur.executemany(query, data_insert)
                    else:
                        print("Data kosong!")
                    
                    # Insert ke tabel summary jika fasilitas aktif
                    if is_active and not result_df.empty:
                        # Kolom yang dibutuhkan untuk summary
                        summary_base = [
                            'namaDebitur', 'npwp', 'ljkKet', 'jenisFasilitasKet', 'kualitas', 'kualitasKet'
                        ]
                        
                        summary_data = result_df[summary_base].copy()
                        summary_data = clean_data_for_sql(summary_data)
                        
                        # Tambah metadata dan nomor laporan
                        for col, val in metadata_columns.items():
                            summary_data[col] = val
                        summary_data['nomorLaporan'] = nomor_laporan
                        
                        # Rename kolom sesuai struktur tabel summary
                        summary_mapping = {
                            'npwp': 'nomorIdentitas',
                            'ljkKet': 'kreditur',
                            'jenisFasilitasKet': 'jenisPembiayaan',
                            'kualitas': 'kodeKolektibilitas',
                            'kualitasKet': 'kolektibilitas'
                        }
                        summary_data = summary_data.rename(columns=summary_mapping)
                        
                        # Tentukan urutan kolom
                        summary_columns = [
                            'periodeData', 'username', 'namaFileUpload', 'uploadDate',
                            'namaDebitur', 'nomorLaporan', 'nomorIdentitas',
                            'kreditur', 'jenisPembiayaan', 'kodeKolektibilitas', 'kolektibilitas'
                        ]
                        
                        summary_df = summary_data[summary_columns]
                        
                        # Buat query insert
                        summary_query = f"""
                            INSERT INTO slik_summary_fasilitas_aktif ({', '.join(summary_columns)})
                            VALUES ({', '.join(['?'] * len(summary_columns))})
                        """
                        summary_tuples = list(summary_df.itertuples(index=False, name=None))
                        
                        if summary_tuples:
                            cur.executemany(summary_query, summary_tuples)
                    
                    # Format data untuk tampilan HTML
                    column_rename_map = {
                        'namaDebitur': 'Nama Debitur/Calon Debitur',
                        'npwp': 'Nomor Identitas',
                        'tglAktaPendirian': 'Tanggal Lahir/Pendirian',
                        'alamat': 'Alamat',
                        'ljkKet': 'Kreditur/Pelapor',
                        'jenisFasilitasKet': 'Jenis Fasilitas',
                        'nominalJumlahKwajibanIDR': 'Oustanding/Baki Debet',
                        'kodeValuta': 'Valuta',
                        'jumlahHariTunggakan': 'Hari Keterlambatan',
                        'kualitas': 'Kode Kolektibilitas Saat ini',
                        'kualitasKet': 'Kolektibilitas Saat ini',
                        'tahunBulan24': 'Periode Pelaporan Terakhir',
                        'urutanFile': 'File ke'
                    }
                    
                    display_df = result_df.rename(columns=column_rename_map)
                    display_df.reset_index(drop=True, inplace=True)
                    
                    # Tambahkan Nomor Laporan jika fasilitas aktif
                    if is_active:
                        display_df.insert(1, 'Nomor Laporan', nomor_laporan)
                    
                    # Tambahkan nomor urut
                    display_df.insert(0, 'No', range(1, len(display_df) + 1))
                    
                    # Hapus kolom metadata
                    display_df = display_df.drop(columns=list(metadata_columns.keys()), errors='ignore')
                    
                    return display_df.to_html(classes="table table-striped", index=False)
                
                # Persiapan data
                uploaded_data_4_dedup = uploaded_data_4.drop_duplicates(subset=['pelapor', 'pelaporKet'])
                combined_data_9 = pd.concat(list_uploaded_data_9, ignore_index=True)
                merged_fLain = combined_data_9.merge(
                    uploaded_data_4_dedup,
                    left_on=['ljk', 'ljkKet'],
                    right_on=['pelapor', 'pelaporKet'],
                    how='left'
                )
                
                if 'npwp' in merged_fLain.columns and 'noIdentitas' in merged_fLain.columns:
                    if 'individual' in data:
                        # Untuk individual: prioritas noIdentitas, fallback ke npwp
                        merged_fLain['npwp'] = merged_fLain.apply(
                            lambda row: row['noIdentitas'] if pd.notna(row['noIdentitas']) and str(row['noIdentitas']).strip() != '' 
                                    else (row['npwp'] if pd.notna(row['npwp']) and str(row['npwp']).strip() != '' else None),
                            axis=1
                        )
                    elif 'perusahaan' in data:
                        # Untuk perusahaan: prioritas npwp, fallback ke noIdentitas
                        merged_fLain['npwp'] = merged_fLain.apply(
                            lambda row: row['npwp'] if pd.notna(row['npwp']) and str(row['npwp']).strip() != '' 
                                    else (row['noIdentitas'] if pd.notna(row['noIdentitas']) and str(row['noIdentitas']).strip() != '' else None),
                            axis=1
                        )
                elif 'noIdentitas' in merged_fLain.columns and 'npwp' not in merged_fLain.columns:
                    # Jika hanya ada noIdentitas, rename ke npwp
                    merged_fLain = merged_fLain.rename(columns={'noIdentitas': 'npwp'})
                
                # Proses fasilitas aktif
                active_fLain = merged_fLain[merged_fLain['kodeKondisi'].isin(['00', '03', '13', '16'])]
                process_other_facility(
                    active_fLain, 
                    'slik_fasilitas_aktif_lainnya',
                    is_active=True
                )
                
                # Proses fasilitas lunas
                closed_fLain = merged_fLain[merged_fLain['kodeKondisi'].isin(['01', '02', '04', '05', '06', '07', '08', '09', '12', '17'])]
                process_other_facility(
                    closed_fLain, 
                    'slik_fasilitas_lunas_lainnya'
                )
                
        if len(list_uploaded_data_10) > 0:
            missing_ljk = [i for i, df in enumerate(list_uploaded_data_10) if 'ljk' not in df.columns]
            if not missing_ljk:
                # 1. Data preparation
                uploaded_data_4_dedup = uploaded_data_4.drop_duplicates(subset=['pelapor', 'pelaporKet'])
                combined_data_10 = pd.concat(list_uploaded_data_10, ignore_index=True)
                merged_fSB = combined_data_10.merge(
                    uploaded_data_4_dedup,
                    left_on=['ljk', 'ljkKet'],
                    right_on=['pelapor', 'pelaporKet'],
                    how='left'
                )
                
                if 'npwp' in merged_fSB.columns and 'noIdentitas' in merged_fSB.columns:
                    if 'individual' in data:
                        # Untuk individual: prioritas noIdentitas, fallback ke npwp
                        merged_fSB['npwp'] = merged_fSB.apply(
                            lambda row: row['noIdentitas'] if pd.notna(row['noIdentitas']) and str(row['noIdentitas']).strip() != '' 
                                    else (row['npwp'] if pd.notna(row['npwp']) and str(row['npwp']).strip() != '' else None),
                            axis=1
                        )
                    elif 'perusahaan' in data:
                        # Untuk perusahaan: prioritas npwp, fallback ke noIdentitas
                        merged_fSB['npwp'] = merged_fSB.apply(
                            lambda row: row['npwp'] if pd.notna(row['npwp']) and str(row['npwp']).strip() != '' 
                                    else (row['noIdentitas'] if pd.notna(row['noIdentitas']) and str(row['noIdentitas']).strip() != '' else None),
                            axis=1
                        )
                elif 'noIdentitas' in merged_fSB.columns and 'npwp' not in merged_fSB.columns:
                    # Jika hanya ada noIdentitas, rename ke npwp
                    merged_fSB = merged_fSB.rename(columns={'noIdentitas': 'npwp'})
                
                # 2. Definisi pemetaan nama kolom untuk output
                column_rename_map = {
                    'namaDebitur': 'Nama Debitur/Calon Debitur',
                    'npwp': 'Nomor Identitas',
                    'tglAktaPendirian': 'Tanggal Pendirian',
                    'tanggalLahir': 'Tanggal Lahir',
                    'alamat': 'Alamat',
                    'ljkKet': 'Kreditur/Pelapor',
                    'jenisSuratBerharga': 'Jenis Surat Berharga',
                    'nilaiPasar': 'Nilai Pasar',
                    'nilaiPerolehan': 'Nilai Perolehan',
                    'nominalSb': 'Outstanding/Baki Debet',
                    'jumlahHariTunggakan': 'Hari Keterlambatan',
                    'kodeValuta': 'Valuta',
                    'kualitas': 'Kode Kolektibilitas Saat ini',
                    'kualitasKet': 'Kolektibilitas Saat ini',
                    'tahunBulan24': 'Periode Pelaporan Terakhir',
                    'urutanFile': 'File ke'
                }
                
                # 3. Pemisahan data berdasarkan kondisi
                active_fSB = merged_fSB[merged_fSB['kondisi'].isin(['00', '03', '13', '16'])]
                closed_fSB = merged_fSB[merged_fSB['kondisi'].isin(['01', '02', '04', '05', '06', '07', '08', '09', '12', '17'])]
                
                # 4. Map kode jenis surat berharga ke deskripsi
                data_df = pd.DataFrame(jenis_surat_berharga)
                kode_to_jenis = data_df.set_index('Kode')['Jenis Surat Berharga'].to_dict()
                
                # 5. Definisi kolom dasar untuk kedua jenis data
                base_columns = [
                    'namaDebitur', 'npwp', 'alamat', 'ljkKet',
                    'jenisSuratBerharga', 'nilaiPasar', 'nilaiPerolehan',
                    'nominalSb', 'jumlahHariTunggakan', 'kualitas',
                    'kualitasKet', 'tahunBulan24', 'urutanFile'
                ]
                
                # Cek apakah kodeValuta harus ditambahkan ke kolom
                columns_with_valuta = base_columns.copy()
                
                if 'tanggalLahir' in merged_fSB.columns:
                    merged_fSB = merged_fSB.rename(columns={'tanggalLahir': 'tglAktaPendirian'})
                if 'tglAktaPendirian' in merged_fSB.columns:
                    columns_with_valuta.insert(2, 'tglAktaPendirian')

                
                if 'kodeValuta' in merged_fSB.columns:
                    columns_with_valuta.insert(9, 'kodeValuta')  # Sisipkan setelah nominalSb
                
                # 6. Function untuk memproses dan menyimpan data
                def process_facility_data(df, columns, table_name):
                    # Map jenis surat berharga
                    df['jenisSuratBerharga'] = df['jenisSuratBerharga'].map(
                        lambda kode: kode_to_jenis.get(kode, kode)
                    )
                    
                    # Ambil data sesuai kolom yang dibutuhkan
                    available_cols = [c for c in columns if c in df.columns]
                    facility_data = df[available_cols].copy()
                    facility_data = clean_data_for_sql(facility_data)
                    
                    # Tambahkan kolom informasi
                    facility_data['periodeData'] = json_posisiDataTerakhir
                    facility_data['username'] = username
                    facility_data['namaFileUpload'] = nama_file
                    facility_data['uploadDate'] = current_datetime
                    
                    # Insert ke database
                    columns_db = ', '.join(facility_data.columns)
                    placeholders = ', '.join(['?'] * len(facility_data.columns))
                    
                    query = f"""
                        INSERT INTO {table_name} ({columns_db})
                        VALUES ({placeholders})
                    """
                    
                    data = list(facility_data.itertuples(index=False, name=None))
                    if data:
                        cur.executemany(query, data)
                    
                    # Format data untuk output HTML
                    output_data = facility_data.rename(columns=column_rename_map)
                    output_data = output_data.drop(columns=['periodeData', 'username', 'namaFileUpload', 'uploadDate'], errors='ignore')
                    
                    # Tambah nomor
                    output_data.reset_index(drop=True, inplace=True)
                    output_data.insert(0, 'No', output_data.index + 1)
                    
                    if table_name == 'slik_fasilitas_aktif_surat_berharga':
                        output_data.insert(1, 'Nomor Laporan', nomor_laporan, allow_duplicates=False)
                        
                    return facility_data
                
                # 7. Proses data untuk active facilities
                active_facility_5 = process_facility_data( 
                    active_fSB, 
                    columns_with_valuta, 
                    'slik_fasilitas_aktif_surat_berharga'
                )
                
                # 8. Proses summary data untuk active facilities
                summary_columns = [
                    'namaDebitur', 'npwp', 'ljkKet', 'jenisSuratBerharga', 'kualitas', 'kualitasKet'
                ]
                
                if active_facility_5.shape[0] > 0:
                    summary_data = active_facility_5[summary_columns].copy()
                    summary_data = clean_data_for_sql(summary_data)
                    summary_data['periodeData'] = json_posisiDataTerakhir
                    summary_data['username'] = username
                    summary_data['namaFileUpload'] = nama_file
                    summary_data['uploadDate'] = current_datetime
                    summary_data['nomorLaporan'] = nomor_laporan
                    
                    # Rename kolom untuk summary
                    summary_rename = {
                        'namaDebitur': 'namaDebitur',
                        'npwp': 'nomorIdentitas',
                        'ljkKet': 'kreditur',
                        'jenisSuratBerharga': 'jenisPembiayaan',
                        'kualitas': 'kodeKolektibilitas',
                        'kualitasKet': 'kolektibilitas'
                    }
                    summary_data = summary_data.rename(columns=summary_rename)
                    
                    # Definisi urutan kolom
                    summary_columns_ordered = [
                        'periodeData', 'username', 'namaFileUpload', 'uploadDate',
                        'namaDebitur', 'nomorLaporan', 'nomorIdentitas',
                        'kreditur', 'jenisPembiayaan', 'kodeKolektibilitas', 'kolektibilitas'
                    ]
                    
                    summary_data = summary_data[summary_columns_ordered]
                    
                    # Insert ke database
                    summary_query = f"""
                        INSERT INTO slik_summary_fasilitas_aktif ({', '.join(summary_columns_ordered)})
                        VALUES ({', '.join(['?'] * len(summary_columns_ordered))})
                    """
                    
                    summary_values = list(summary_data.itertuples(index=False, name=None))
                    if summary_values:
                        cur.executemany(summary_query, summary_values)
                
                # 9. Proses data untuk closed facilities
                closed_facility_5 = process_facility_data(
                    closed_fSB, 
                    columns_with_valuta, 
                    'slik_fasilitas_lunas_surat_berharga'
                )
            
        # Update progress to 100% (completed)
        conn.commit()
        if task_id in task_progress:
            task_progress[task_id].update({
                'progress': 100,
                'status': 'completed',
                'completed_at': datetime.now().isoformat(),
                'message': 'Upload berhasil diproses'
            })
        
        # PERBAIKAN: Store result dengan flag untuk stop polling
        task_results[task_id] = {
            "status": "success",
            "completed": True,
            "redirect_url": "/upload-success",  # atau URL tujuan setelah selesai
            "data": list_table_data,
            "message": "File berhasil diproses"
        }
        
        # Log completion
        app.logger.info(f"Task {task_id} completed successfully")
        
        # Return all processed data
        return {
            "table_data": table_data,
            "list_table_data": list_table_data,
        }
    except Exception as e:
        # PERBAIKAN: Update task progress untuk error state
        if task_id in task_progress:
            task_progress[task_id].update({
                'progress': 0,
                'status': 'error',
                'completed_at': datetime.now().isoformat(),
                'error_message': str(e),
                'message': f'Error processing file: {str(e)}'
            })
        
        # PERBAIKAN: Store error result dengan flag untuk stop polling
        task_results[task_id] = {
            "status": "error",
            "completed": True,
            "error": True,
            "error_type": "system_error",
            "message": f"Error processing file: {str(e)}",
            "redirect_url": "/upload-big-size"
        }
        
        # Rollback database changes
        if conn:
            conn.rollback()
        
        # Log error
        app.logger.error(f"[{task_id}] Error processing file: {str(e)}")
        
        # Return error response
        return {
            "error": True,
            "error_type": "system_error", 
            "message": f"Error processing file: {str(e)}",
            "redirect_url": "/upload-big-size"
        }
    finally:
        # PERBAIKAN: Pastikan connection ditutup
        if conn:
            conn.close()

# Daftar tabel yang akan diekspor
tables = [
    "slik_summary_fasilitas_aktif",
    "slik_fasilitas_aktif_kredit_pembiayaan",
    "slik_fasilitas_aktif_bank_garansi",
    "slik_fasilitas_aktif_lainnya",
    "slik_fasilitas_aktif_lc",
    "slik_fasilitas_aktif_surat_berharga",
    "slik_fasilitas_lunas_kredit_pembiayaan",
    "slik_fasilitas_lunas_bank_garansi",
    "slik_fasilitas_lunas_lainnya",
    "slik_fasilitas_lunas_lc",
    "slik_fasilitas_lunas_surat_berharga"
]

# Perbaikan untuk fungsi get_progress_status dengan timeout protection
@app.route('/progress-status/<task_id>')
def get_progress_status(task_id):
    """
    Get progress status dengan ultra-fast response dan minimal locking.
    FIXES:
    - Reduced lock time to < 1ms
    - Simple data structure
    - No complex validations
    - Fast-fail on missing task
    """
    try:
        # PERBAIKAN: Single lock acquisition dengan minimal processing
        response_data = None
        
        with task_progress_lock:
            if task_id in task_progress:
                # Quick copy of essential data only
                progress_data = task_progress[task_id]
                response_data = {
                    'progress': int(progress_data.get('progress', 0)),
                    'status': str(progress_data.get('status', 'processing')),
                    'timestamp': float(progress_data.get('timestamp', time.time())),
                    'upload_session': progress_data.get('upload_session'),
                    'task_id': task_id,
                    'message': progress_data.get('message', ''),
                    'progress_bars': progress_data.get('progress_bars', {}),
                    'metadata': progress_data.get('temp_metadata', {})
                }
        
        # Process response outside lock
        if response_data:
            progress = response_data['progress']
            status = response_data['status']
            
            # Add completion flags
            if status == 'completed':
                response_data.update({
                    'completed': True,
                    'should_redirect': True,
                    'redirect_url': '/upload-success',
                    'final_message': 'Upload berhasil diproses',
                    'error': False
                })
            elif status == 'error':
                response_data.update({
                    'completed': True,
                    'should_redirect': True,
                    'redirect_url': '/upload-big-size',
                    'error': True,
                    'error_type': response_data.get('error_type', 'general_error')
                })
            
            return jsonify(response_data)
        
        # Check in task_results (fallback)
        if task_id in task_results:
            result = task_results[task_id]
            
            response = {
                'progress': 100,
                'timestamp': float(time.time()),
                'task_id': task_id
            }
            
            if result.get('status') == 'success':
                response.update({
                    'status': 'completed',
                    'completed': True,
                    'should_redirect': True,
                    'redirect_url': result.get('redirect_url', '/upload-success'),
                    'message': result.get('message', 'Upload berhasil diproses'),
                    'error': False
                })
            else:
                response.update({
                    'status': 'error',
                    'completed': True,
                    'should_redirect': True,
                    'redirect_url': result.get('redirect_url', '/upload-big-size'),
                    'error': True,
                    'message': result.get('message', 'Task failed'),
                    'error_type': result.get('error_type', 'general_error')
                })
            
            return jsonify(response)
        
        # Task not found
        return jsonify({
            'progress': 0,
            'status': 'not_found',
            'message': 'Task not found',
            'timestamp': float(time.time()),
            'error': True,
            'task_id': task_id
        }), 404
    
    except Exception as e:
        app.logger.error(f"❌ Error in get_progress_status: {e}")
        return jsonify({
            'progress': 0,
            'status': 'error',
            'message': 'Server error',
            'error': True,
            'task_id': task_id,
            'timestamp': float(time.time())
        }), 500
        
# Fungsi untuk mengekspor data tabel ke Excel (versi sederhana)
def export_to_excel(periodeData, username, namaFileUpload, uploadDate):
        
    # Format tanggal untuk query
    uploadDate_str = uploadDate
    if 'T' in uploadDate:
        uploadDate_str = uploadDate.replace('T', ' ')
    
    # Definisikan tabel dan nama sheet
    tables = [
        'slik_summary_fasilitas_aktif',
        'slik_fasilitas_aktif_kredit_pembiayaan',
        'slik_fasilitas_aktif_bank_garansi',
        'slik_fasilitas_aktif_lainnya',
        'slik_fasilitas_aktif_lc',
        'slik_fasilitas_aktif_surat_berharga',
        'slik_fasilitas_lunas_kredit_pembiayaan',
        'slik_fasilitas_lunas_bank_garansi',
        'slik_fasilitas_lunas_lainnya',
        'slik_fasilitas_lunas_lc',
        'slik_fasilitas_lunas_surat_berharga'
    ]
    
    sheet_name_mapping = {
        'slik_summary_fasilitas_aktif': 'Rangkuman Fasilitas Aktif',
        'slik_fasilitas_aktif_kredit_pembiayaan': 'Fasilitas Aktif Kredit Pembiayaan',
        'slik_fasilitas_aktif_bank_garansi': 'Fasilitas Aktif Bank Garansi',
        'slik_fasilitas_aktif_lainnya': 'Fasilitas Aktif Lainnya',
        'slik_fasilitas_aktif_lc': 'Fasilitas Aktif LC',
        'slik_fasilitas_aktif_surat_berharga': 'Fasilitas Aktif Surat Berharga',
        'slik_fasilitas_lunas_kredit_pembiayaan': 'Fasilitas Lunas Kredit Pembiayaan',
        'slik_fasilitas_lunas_bank_garansi': 'Fasilitas Lunas Bank Garansi',
        'slik_fasilitas_lunas_lainnya': 'Fasilitas Lunas Lainnya',
        'slik_fasilitas_lunas_lc': 'Fasilitas Lunas LC',
        'slik_fasilitas_lunas_surat_berharga': 'Fasilitas Lunas Surat Berharga'
    }
    
    # Mapping kolom yang umum digunakan di semua sheet
    def get_column_mapping(sheet_name):
        # Base mapping yang berlaku untuk semua sheet
        base_mapping = {
            'No': 'No',
            'namaDebitur': 'Nama Debitur/Calon Debitur',
            'npwp': 'Nomor Identitas',
            'nomorIdentitas': 'Nomor Identitas',
            'alamat': 'Alamat',
            'ljkKet': 'Kreditur/Pelapor',
            'kreditur': 'Kreditur/Pelapor',
            'plafon': 'Plafon',
            'kualitas': 'Kode Kolektibilitas Saat ini',
            'kodeKolektibilitas': 'Kode Kolektibilitas Saat ini',
            'kualitasKet': 'Kolektibilitas Saat ini',
            'kolektibilitas': 'Kolektibilitas Saat ini',
            'tahunBulan24': 'Periode Pelaporan Terakhir',
            'urutanFile': 'File ke',
            'tglAktaPendirian': 'Tanggal Pendirian',
            'tanggalLahir': 'Tanggal Lahir',
            'kodeValuta': 'Valuta',
            'valuta': 'Valuta',
            'valutaKode': 'Valuta',
            'jumlahHariTunggakan': 'Hari Keterlambatan'
        }
        
        # Mapping spesifik berdasarkan jenis sheet
        if 'Rangkuman' in sheet_name:
            specific_mapping = {
                'nomorLaporan': 'Nomor Laporan',
                'jenisPembiayaan': 'Jenis Kredit/Pembiayaan'
            }
        elif 'Kredit Pembiaya' in sheet_name:
            specific_mapping = {
                'jenisKreditPembiayaanKet': 'Jenis Kredit/Pembiayaan',
                'jenisPenggunaanKet': 'Jenis Penggunaan',
                'bakiDebet': 'Oustanding/Baki Debet',
                'tunggakanPokok': 'Tunggakan Pokok',
                'tunggakanBunga': 'Tunggakan Bunga',
                'denda': 'Denda'
            }
        elif 'LC' in sheet_name:
            specific_mapping = {
                'jenisLcKet': 'Jenis L/C',
                'tujuanLcKet': 'Tujuan L/C',
                'nominalLc': 'Oustanding/Baki Debet',
                'tanggalWanPrestasi': 'Tanggal Wan prestasi'
            }
        elif 'Bank Garansi' in sheet_name:
            specific_mapping = {
                'jenisGaransiKet': 'Jenis Garansi',
                'tujuanGaransiKet': 'Tujuan Garansi',
                'nominalBg': 'Oustanding/Baki Debet',
                'tanggalWanPrestasi': 'Tanggal Wan Prestasi'
            }
        elif 'Lainnya' in sheet_name:
            specific_mapping = {
                'jenisFasilitasKet': 'Jenis Fasilitas',
                'nominalJumlahKwajibanIDR': 'Oustanding/Baki Debet'
            }
        elif 'Surat Berharga' in sheet_name:
            specific_mapping = {
                'jenisSuratBerharga': 'Jenis Surat Berharga',
                'nilaiPasar': 'Nilai Pasar',
                'nilaiPerolehan': 'Nilai Perolehan',
                'nominalSb': 'Outstanding/Baki Debet'
            }
        else:
            specific_mapping = {}
        
        # Gabungkan base mapping dengan specific mapping
        return {**base_mapping, **specific_mapping}
    
    # Buat Excel workbook
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
        
        # Format untuk highlight baris kuning
        highlight_format = workbook.add_format({'bg_color': '#FFFF00', 'border': 1})
        normal_format = workbook.add_format({'border': 1})
        header_format = workbook.add_format({'bold': True, 'border': 1})
        
        data_found = False
        print(f"Periode: {periodeData} username: {username} and file: {namaFileUpload}")
        for table in tables:
            try:
                query = f"""
                    SELECT * FROM {table}
                    WHERE periodeData = ? AND namaFileUpload = ?
                """
                cur.execute(query, (periodeData, namaFileUpload))

                columns = [column[0] for column in cur.description]
                rows = cur.fetchall()
                df = pd.DataFrame.from_records(rows, columns=columns)
                print(f"Found {len(df)} rows for table {table}")
                
                # Drop kolom yang tidak perlu
                drop_cols = ['periodeData', 'username', 'namaFileUpload', 'uploadDate', 'id']
                df.drop(columns=[col for col in drop_cols if col in df.columns], errors='ignore', inplace=True)

                # Validasi khusus untuk kolom tanggal - drop jika semua nilai kosong
                date_cols_to_check = ['tglAktaPendirian', 'tanggalLahir']
                for col in date_cols_to_check:
                    if col in df.columns:
                        # Cek apakah semua nilai dalam kolom kosong (None, NaN, atau string kosong)
                        if df[col].isna().all() or (df[col].astype(str).str.strip() == '').all():
                            print(f"Dropping column '{col}' - all values are empty")
                            df.drop(columns=[col], inplace=True)
                
                # Tambahkan kolom No
                df.insert(0, 'No', range(1, len(df) + 1))

                sheet_name = sheet_name_mapping.get(table, table)[:31]
                
                # Rename kolom
                mapping = get_column_mapping(sheet_name)
                rename_dict = {old: new for old, new in mapping.items() if old in df.columns}
                df.rename(columns=rename_dict, inplace=True)
                
                # Buat worksheet
                worksheet = workbook.add_worksheet(sheet_name)
                
                # Tulis header
                for col_idx, col_name in enumerate(df.columns):
                    worksheet.write(0, col_idx, col_name, header_format)
                
                # Cari kolom untuk highlight
                highlight_col = None
                for i, col_name in enumerate(df.columns):
                    if col_name == 'Kode Kolektibilitas Saat ini':
                        highlight_col = i
                        break
                
                # Tulis data dengan format kondisional
                for row_idx in range(len(df)):
                    need_highlight = False
                    if highlight_col is not None:
                        value = df.iloc[row_idx, highlight_col]
                        if pd.notna(value) and str(value) != '1':
                            need_highlight = True
                    
                    format_to_use = highlight_format if need_highlight else normal_format
                    excel_row = row_idx + 1
                    for col_idx in range(len(df.columns)):
                        worksheet.write(excel_row, col_idx, df.iloc[row_idx, col_idx], format_to_use)
                
                data_found = True
                print(f"Wrote data to sheet {sheet_name}")
            except Exception as e:
                print(f"Error on table {table}: {str(e)}")
                continue
        
        # Jika tidak ada data
        if not data_found:
            pd.DataFrame({'Message': ['No data found for the specified criteria']}).to_excel(
                writer, sheet_name='No Data', index=False)
    
    output.seek(0)
    return output

@app.route('/upload-big-size', methods=['GET', 'POST'])
def upload_big_size_file():
    if 'username' not in session:
        flash("Please log in first.")
        return redirect(url_for('login'))

    role_access = session.get('role_access')
    fullname = session.get('fullname')
    username = session.get('username')
    report_access = session.get('report_access')

    if request.method == 'GET':
        show_alert = session.pop('upload_done', False)
        show_processing_alert = session.pop('show_processing_alert', False)
        
        # HANYA collect active task IDs untuk user ini
        valid_task_ids = []
        
        try:
            # Quick snapshot dengan minimal lock
            with task_id_registry_lock:
                # Get ALL active tasks untuk user ini dari registry
                for tid, info in task_id_registry.items():
                    if (info.get('username') == username and 
                        info.get('status') == 'active'):
                        valid_task_ids.append(tid)
            
            # Add current task_id dari session jika ada
            session_task_id = session.get('task_id')
            if session_task_id and session_task_id not in valid_task_ids:
                # Validasi task masih exist di registry
                with task_id_registry_lock:
                    if session_task_id in task_id_registry:
                        if task_id_registry[session_task_id].get('status') == 'active':
                            valid_task_ids.append(session_task_id)
                    else:
                        # Task sudah tidak ada, clear dari session
                        session.pop('task_id', None)
                        session_task_id = None
            
            app.logger.info(
                f"📋 GET /upload-big-size - User: {username}, "
                f"Active tasks: {len(valid_task_ids)}"
            )
            
        except Exception as e:
            app.logger.error(f"❌ Error collecting active tasks: {e}")
            valid_task_ids = []
            session_task_id = None

        return render_template(
            'upload_big_size.html',
            show_alert=show_alert,
            show_processing_alert=show_processing_alert,
            flags=FLAGS,
            role_access=role_access,
            fullname=fullname,
            report_access=report_access,
            task_progress={},  # KOSONG - akan dimuat via AJAX
            task_id=session_task_id,  # Current task dari session
            tasksFromBackend=valid_task_ids  # ALL active tasks untuk polling
        )

    elif request.method == 'POST':
        try:
            flag = request.form.get('flag')
            nama_file = request.form.get('nama_file')
            uploaded_files = request.files.getlist('file')

            if not uploaded_files or not any(f.filename for f in uploaded_files):
                error_msg = 'No files selected for upload.'
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'error': error_msg}), 400
                flash(error_msg)
                return redirect(url_for('upload_big_size_file'))
            
            if not flag or not nama_file:
                error_msg = 'Semua field harus diisi'
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'error': error_msg}), 400
                flash(error_msg, 'error')
                return redirect(url_for('upload_big_size_file'))
            
            # Check duplicate filename
            if check_filename_exists(nama_file):
                error_msg = 'Nama File sudah ada, gunakan nama lain'
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'error': error_msg}), 400
                flash(error_msg, 'error')
                return redirect(url_for('upload_big_size_file'))
            
            total_file_size = 0
            for f in uploaded_files:
                if f and f.filename:
                    if not is_txt_file(f.filename):
                        error_msg = f"File '{f.filename}' bukan file .txt"
                        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return jsonify({'success': False, 'error': error_msg}), 400
                        return f'<script>alert("{error_msg}"); window.location.href = "/upload-big-size";</script>'
                    
                    f.seek(0, 2)
                    total_file_size += f.tell()
                    f.seek(0)
            
            if total_file_size > MAX_FILE_BIG_SIZE:
                error_msg = 'Total file terlalu besar. Maksimum 200MB!'
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'error': error_msg}), 400
                return f'<script>alert("{error_msg}"); window.location.href = "/upload-big-size";</script>'
            
            task_id = generate_unique_task_id(username, nama_file)
            
            # Get upload session dari registry (sudah di-set oleh generate_unique_task_id)
            with task_id_registry_lock:
                upload_session = task_id_registry[task_id]['upload_session']
            
            current_datetime = datetime.now()
            
            # Read files into memory
            temp_files = []
            for file in uploaded_files:
                if file and file.filename:
                    file_content = file.read()
                    temp_files.append({
                        'filename': file.filename,
                        'content': file_content,
                        'mimetype': file.content_type or 'text/plain'
                    })
            
            # Prepare user info
            user_info = {
                "username": username,
                "role_access": role_access,
                "fullname": fullname,
                "flag": flag,
                "nama_file": nama_file
            }
            
            with task_progress_lock:
                task_progress[task_id] = {
                    'progress': 0,
                    'status': 'processing',
                    'timestamp': time.time(),
                    'created_at': time.time(),
                    'key': f"{username}_{nama_file}_{flag}",
                    'upload_session': upload_session,
                    'message': 'Mempersiapkan upload...',
                    'temp_metadata': {
                        'username': username,
                        'namaFileUpload': nama_file,
                        'periodeData': '-',
                        'uploadDate': current_datetime.strftime('%d %B %Y, %H:%M'),
                        'task_id': task_id,
                        'fullname': fullname
                    },
                    'progress_bars': {
                        'file_processing': 0,
                        'db_processing': 0
                    }
                }
            
            try:
                task_queue.put_nowait((task_id, temp_files, user_info))
                
                app.logger.info(
                    f"✅ Task queued - ID: {task_id}, Session: {upload_session}, "
                    f"User: {username}, File: {nama_file}, Size: {total_file_size} bytes"
                )
                
            except queue.Full:
                # Cleanup on queue full
                with task_progress_lock:
                    task_progress.pop(task_id, None)
                mark_task_error(task_id)
                
                error_msg = 'Antrean upload penuh. Silakan coba lagi nanti.'
                
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'error': error_msg}), 400
                
                flash(error_msg, "error")
                return redirect(url_for('upload_big_size_file'))
            
            session['task_id'] = task_id
            session['upload_session'] = upload_session
            session['upload_done'] = False
            session['show_processing_alert'] = True
            session['temp_filename'] = nama_file
            
            app.logger.info(
                f"✅ Upload initiated successfully - "
                f"Task: {task_id}, Session: {upload_session}"
            )
            
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'success': True, 
                    'task_id': task_id, 
                    'upload_session': upload_session,
                    'message': 'Upload started successfully'
                }), 200
            
            return redirect(url_for('upload_big_size_file'))
        
        except Exception as e:
            app.logger.error(f"❌ Error during upload: {e}", exc_info=True)
            
            # Cleanup task progress if exists
            if 'task_id' in locals():
                with task_progress_lock:
                    task_progress.pop(task_id, None)
                mark_task_error(task_id)
            
            error_msg = f'Error during upload: {str(e)}'
            
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'error': error_msg}), 500
            
            flash(error_msg, 'error')
            return redirect(url_for('upload_big_size_file'))
        
cleanup_old_tasks()

@app.route('/api/upload-data')
def api_upload_data():
    """
    Endpoint untuk load data upload dengan:
    - NO DUPLICATE rows
    - Progress persistence
    - Fast queries
    
    FIXES:
    - Gunakan Set untuk track filenames yang sudah ditampilkan
    - Skip database rows jika file sedang di-process
    - Return task_id dalam response untuk frontend tracking
    """
    MAX_QUERY_TIMEOUT = 5
    start_time = time.time()
    
    try:
        # Get pagination parameters
        draw = int(request.args.get('draw', 1))
        start = int(request.args.get('start', 0))
        length = int(request.args.get('length', 10))
        search_value = request.args.get('search[value]', '')
        order_column_index = int(request.args.get('order[0][column]', 2))
        order_direction = request.args.get('order[0][dir]', 'desc')
        
        column_map = {
            0: 'namaFileUpload',
            1: 'periodeData',
            2: 'uploadDate'
        }
        order_column = column_map.get(order_column_index, 'uploadDate')

        in_progress_data = []
        in_progress_filenames = set()  # Track filenames untuk avoid duplikat
        in_progress_tasks = {}  # Track task details untuk frontend
        
        try:
            with task_progress_lock:
                # Get active tasks dari registry
                active_task_ids = set()
                with task_id_registry_lock:
                    for tid, info in task_id_registry.items():
                        if info.get('status') == 'active':
                            active_task_ids.add(tid)
                
                # Collect progress untuk active tasks
                snapshot = []
                for tid in active_task_ids:
                    if tid in task_progress:
                        snapshot.append((tid, task_progress[tid].copy()))
            
            # Process snapshot
            for tid, v in snapshot:
                status = v.get('status', 'processing')
                
                if status == 'processing' and 'temp_metadata' in v:
                    meta = v['temp_metadata'].copy()
                    progress = int(v.get('progress', 0))
                    upload_session = v.get('upload_session', '')
                    message = v.get('message', '')
                    
                    filename = meta.get('namaFileUpload', '')
                    
                    # CRITICAL: Track filename untuk prevent duplikasi
                    if filename:
                        in_progress_filenames.add(filename)
                    
                    # CRITICAL: Store task info untuk frontend
                    in_progress_tasks[tid] = {
                        'filename': filename,
                        'upload_session': upload_session,
                        'progress': progress
                    }
                    
                    # Generate HTML dengan task_id attribute
                    animated_class = 'progress-bar-animated' if progress < 100 else ''
                    progress_text = f"{progress}%"
                    if message and progress < 100:
                        progress_text = f"{progress}% - {message}"
                    
                    # PERBAIKAN: Gunakan format ID yang konsisten
                    progress_bar_id = f"progress-bar-{tid}"
                    
                    progress_html = f'''
                    <div class="progress" style="height:18px;">
                        <div id="{progress_bar_id}" 
                             class="progress-bar progress-bar-striped {animated_class}" 
                             role="progressbar" 
                             style="width: {progress}%;"
                             data-task-id="{tid}"
                             data-upload-session="{upload_session}"
                             data-filename="{filename}">
                            {progress_text}
                        </div>
                    </div>
                    '''
                    
                    download_html = f'''
                    <button id="download-btn-{tid}" 
                            class="btn btn-secondary btn-sm" 
                            data-task-id="{tid}"
                            disabled>
                        <i class="fas fa-spinner fa-spin"></i> {progress}%
                    </button>
                    '''
                    
                    in_progress_data.append([
                        meta.get('namaFileUpload', '-'),
                        meta.get('periodeData', '-'),
                        meta.get('uploadDate', '-'),
                        progress_html,
                        download_html
                    ])
            
            app.logger.info(
                f"📊 Found {len(in_progress_data)} in-progress tasks: "
                f"{list(in_progress_filenames)}"
            )
            
        except Exception as cache_err:
            app.logger.error(f"❌ Error collecting in-progress: {cache_err}")
        
        # Check timeout
        elapsed = time.time() - start_time
        if elapsed > MAX_QUERY_TIMEOUT * 0.2:
            app.logger.warning(f"⏱️ Slow start, returning in-progress only")
            return jsonify({
                "draw": draw,
                "recordsTotal": len(in_progress_data),
                "recordsFiltered": len(in_progress_data),
                "data": in_progress_data,
                "inProgressTasks": list(in_progress_tasks.keys())  # Send task IDs
            })
        
        conn = None
        cursor = None
        db_data = []
        total_records = 0
        
        try:
            conn = get_db_connection()
            conn.timeout = MAX_QUERY_TIMEOUT
            cursor = conn.cursor()
            
            base_query = """
                SELECT periodeData, namaFileUpload, uploadFolderPath, username, 
                       fullname, uploadDate
                FROM slik_uploader WITH (NOLOCK)
            """
            
            where_clause = ""
            params = []
            if search_value:
                where_clause = """ WHERE 
                    namaFileUpload LIKE ? OR 
                    periodeData LIKE ? OR 
                    username LIKE ? OR 
                    fullname LIKE ?
                """
                search_param = f'%{search_value}%'
                params = [search_param] * 4
            
            # Count
            count_query = f"SELECT COUNT(*) FROM slik_uploader WITH (NOLOCK) {where_clause}"
            cursor.execute(count_query, params)
            total_records = cursor.fetchone()[0]
            
            # Check timeout
            elapsed = time.time() - start_time
            if elapsed > MAX_QUERY_TIMEOUT * 0.6:
                raise TimeoutError(f"Query timeout: {elapsed:.2f}s")
            
            # Get data
            query = f"""
                {base_query}
                {where_clause}
                ORDER BY {order_column} {order_direction.upper()}
                OFFSET ? ROWS
                FETCH NEXT ? ROWS ONLY
            """
            params_query = params.copy()
            params_query.extend([start, length])
            
            cursor.execute(query, params_query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            
            # Process rows dengan SKIP duplikasi
            skipped_count = 0
            for row in rows:
                item = dict(zip(columns, row))
                filename = item.get('namaFileUpload', '')
                if filename in in_progress_filenames:
                    skipped_count += 1
                    app.logger.debug(
                        f"⏭️ Skipping '{filename}' - already in progress"
                    )
                    continue
                
                # Parse date
                try:
                    upload_date_obj = parser.parse(str(item.get("uploadDate")))
                    upload_date = upload_date_obj.strftime("%d %B %Y, %H:%M")
                    upload_date_raw = item.get("uploadDate")
                except:
                    upload_date = "N/A"
                    upload_date_raw = ""
                
                # HTML untuk completed items
                progress_html = '''
                <div class="progress" style="height:18px;">
                    <div class="progress-bar" role="progressbar" style="width: 100%;">
                        100%
                    </div>
                </div>
                '''
                
                download_html = f'''
                <div class="btn-group">
                    <a href="{url_for('download_big_size', 
                                      periodeData=item['periodeData'], 
                                      username=item['username'], 
                                      namaFileUpload=item['namaFileUpload'], 
                                      uploadDate=str(upload_date_raw))}" 
                       class="btn btn-outline-success btn-sm" title="Download Excel">
                        Excel <i class="fas fa-file-excel"></i>
                    </a>
                    <a href="{url_for('download_upload_zip', 
                                      periodeData=item['periodeData'], 
                                      username=item['username'], 
                                      namaFileUpload=item['namaFileUpload'], 
                                      uploadDate=str(upload_date_raw))}" 
                       class="btn btn-outline-primary btn-sm" title="Download ZIP">
                        ZIP <i class="fas fa-file-archive"></i>
                    </a>
                </div>
                '''
                
                db_data.append([
                    item['namaFileUpload'],
                    item['periodeData'],
                    upload_date,
                    progress_html,
                    download_html
                ])
            
            if skipped_count > 0:
                app.logger.info(f"⏭️ Skipped {skipped_count} duplicate rows")
        
        except TimeoutError as te:
            app.logger.warning(f"⏱️ Database timeout: {te}")
            return jsonify({
                "draw": draw,
                "recordsTotal": len(in_progress_data),
                "recordsFiltered": len(in_progress_data),
                "data": in_progress_data,
                "inProgressTasks": list(in_progress_tasks.keys()),
                "warning": "Database timeout"
            })
        
        except Exception as db_err:
            app.logger.error(f"❌ Database error: {db_err}")
            return jsonify({
                "draw": draw,
                "recordsTotal": len(in_progress_data),
                "recordsFiltered": len(in_progress_data),
                "data": in_progress_data,
                "inProgressTasks": list(in_progress_tasks.keys()),
                "error": "Database error"
            })
        
        finally:
            if cursor:
                try:
                    cursor.close()
                except:
                    pass
            if conn:
                try:
                    conn.close()
                except:
                    pass
        
        all_data = in_progress_data + db_data
        
        elapsed = time.time() - start_time
        if elapsed > 2.0:
            app.logger.warning(f"⚠️ API took {elapsed:.2f}s")
        
        app.logger.info(
            f"✅ API response: {len(in_progress_data)} in-progress + "
            f"{len(db_data)} completed = {len(all_data)} total"
        )
        
        return jsonify({
            "draw": draw,
            "recordsTotal": total_records + len(in_progress_data),
            "recordsFiltered": total_records + len(in_progress_data),
            "data": all_data,
            "inProgressTasks": list(in_progress_tasks.keys())  # CRITICAL untuk resume polling
        })
    
    except Exception as e:
        app.logger.error(f"❌ Fatal error: {e}", exc_info=True)
        return jsonify({
            "draw": draw if 'draw' in locals() else 1,
            "recordsTotal": 0,
            "recordsFiltered": 0,
            "data": [],
            "error": str(e)
        }), 500
                
@app.route('/download-big-size', methods=['GET'])
def download_big_size():
    try:
        periodeData = urllib.parse.unquote(request.args.get("periodeData", ""))
        username = session.get('username')
        namaFileUpload = urllib.parse.unquote(request.args.get("namaFileUpload", ""))
        uploadDate = urllib.parse.unquote(request.args.get("uploadDate", ""))
        downloadType = "Output Excel"
        downloadDate = datetime.now()

        conn = None
        cursor = None

        try:
            dt_obj = try_parse_upload_date(uploadDate)
            uploadDate = dt_obj.strftime("%Y-%m-%d %H:%M:%S")

            conn = get_db_connection()
            cursor = conn.cursor()

            query = """
                INSERT INTO slik_download_logging (periodeData, namaFileUpload, downloadType, username, downloadDate)
                VALUES (?, ?, ?, ?, ?)
            """
            cursor.execute(query, (periodeData, namaFileUpload, downloadType, username, downloadDate))
            conn.commit()

        except Exception as e:
            print(f"Error saving to DB: {e}")
            traceback.print_exc()
        finally:
            if cursor: cursor.close()
            if conn: conn.close()

        print(f"Downloading Excel with parameters: {periodeData}, {username}, {namaFileUpload}, {uploadDate}")
        excel_file = export_to_excel(periodeData, username, namaFileUpload, uploadDate)

        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f"{periodeData}_{namaFileUpload}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"Error in download_excel: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/handle-upload-error/<task_id>')
def handle_upload_error(task_id):
    """Handle upload errors and redirect appropriately"""
    if task_id in task_results:
        result = task_results[task_id]
        if result.get('status') == 'error':
            error_type = result.get('error_type', 'general_error')
            error_message = result.get('error', 'Unknown error')
            redirect_url = result.get('redirect_url', '/upload-big-size')
            
            # Clean up task results after handling
            del task_results[task_id]
            if task_id in task_progress:
                del task_progress[task_id]
            
            return f'''
                <script>
                    alert("{error_message}");
                    window.location.href = "{redirect_url}";
                </script>
            '''
    
    return redirect(url_for('upload_big_size_file'))

def convert_bulan_sql(column_name: str) -> str:
    """
    Mengonversi nama bulan Indonesia menjadi Inggris untuk kebutuhan filter periode.
    """
    return f"""
        CASE
            WHEN LEFT({column_name}, CHARINDEX(' ', {column_name}) - 1) = 'Januari' THEN REPLACE({column_name}, 'Januari', 'January')
            WHEN LEFT({column_name}, CHARINDEX(' ', {column_name}) - 1) = 'Februari' THEN REPLACE({column_name}, 'Februari', 'February')
            WHEN LEFT({column_name}, CHARINDEX(' ', {column_name}) - 1) = 'Maret' THEN REPLACE({column_name}, 'Maret', 'March')
            WHEN LEFT({column_name}, CHARINDEX(' ', {column_name}) - 1) = 'April' THEN REPLACE({column_name}, 'April', 'April')
            WHEN LEFT({column_name}, CHARINDEX(' ', {column_name}) - 1) = 'Mei' THEN REPLACE({column_name}, 'Mei', 'May')
            WHEN LEFT({column_name}, CHARINDEX(' ', {column_name}) - 1) = 'Juni' THEN REPLACE({column_name}, 'Juni', 'June')
            WHEN LEFT({column_name}, CHARINDEX(' ', {column_name}) - 1) = 'Juli' THEN REPLACE({column_name}, 'Juli', 'July')
            WHEN LEFT({column_name}, CHARINDEX(' ', {column_name}) - 1) = 'Agustus' THEN REPLACE({column_name}, 'Agustus', 'August')
            WHEN LEFT({column_name}, CHARINDEX(' ', {column_name}) - 1) = 'September' THEN REPLACE({column_name}, 'September', 'September')
            WHEN LEFT({column_name}, CHARINDEX(' ', {column_name}) - 1) = 'Oktober' THEN REPLACE({column_name}, 'Oktober', 'October')
            WHEN LEFT({column_name}, CHARINDEX(' ', {column_name}) - 1) = 'November' THEN REPLACE({column_name}, 'November', 'November')
            WHEN LEFT({column_name}, CHARINDEX(' ', {column_name}) - 1) = 'Desember' THEN REPLACE({column_name}, 'Desember', 'December')
        END
    """

# Daftar Fasilitas Debitur SLIK OJK
@app.route('/daftar-fasilitas-debitur', methods=['GET'])
def daftar_fasilitas_debitur_page():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    return render_template(
        'daftar_fasilitas_debitur.html',
        username=session.get('username'),
        fullname=session.get('fullname'),
        role_access=session.get('role_access'),
        report_access=session.get('report_access')
    )
            
@app.route('/api/data-fasilitas-debitur', methods=['GET'])
def api_fasilitas_debitur_page():
    """
    Endpoint untuk data fasilitas debitur dengan filter periode (YYYYMM), pagination, dan limit.
    Query params: periode (format: YYYYMM), page, page_size
    """
    try:
        # Ambil parameter dari request
        periode = request.args.get('periode')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 50))
        offset = (page - 1) * page_size

        conn = get_db_connection()
        cursor = conn.cursor()

        # --- WHERE dasar ---
        base_where = """
            WHERE A.bentukBU IS NOT NULL
              AND LEN(A.npwp) >= 5
              AND B.jenisKreditPembiayaan IS NOT NULL
        """
        params = []

        # --- Filter berdasarkan periode (YYYYMM) ---
        if periode:
            periode_expr = f"LEFT(CONVERT(CHAR(8), CONVERT(DATE, '01 ' + {convert_bulan_sql('A.periodeData')}, 107), 112), 6)"
            base_where += f" AND {periode_expr} = ?"
            params.append(periode)

        # --- Query utama (base) ---
        base_query = f"""
            FROM [dbo].[slik_data_pokok_debitur] A
            LEFT JOIN [dbo].[slik_fasilitas_kredit_pembiayaan] B
                ON A.namaFileUpload = B.namaFileUpload 
                AND A.periodeData = B.periodeData 
                AND A.pelaporKet = B.ljkKet
            {base_where}
        """

        # --- Query hitung total ---
        count_query = f"SELECT COUNT(*) {base_query}"

        # Jalankan count
        cursor.execute(count_query, params)
        total_records = cursor.fetchone()[0]

        # --- Query ambil data (pagination) ---
        data_query = f"""
            SELECT
                CONVERT(CHAR(8),
                    EOMONTH(CONVERT(DATE, '01 ' + {convert_bulan_sql('A.periodeData')}, 107)), 112
                ) AS [Periode Data],
                A.npwp AS [ID Debitur/ NPWP],
                A.namaDebitur AS [Nama Debitur],
                CONCAT(B.jenisKreditPembiayaan, ' - ', B.jenisKreditPembiayaanKet) AS [Jenis Kredit / Pembiayaan],
                CONCAT(B.akadKreditPembiayaan, ' - ', B.akadKreditPembiayaanKet) AS [Akad Kredit / Pembiayaan],
                A.pelaporKet AS [Pelapor],
                B.bakiDebet AS [Baki Debet],
                CONCAT(B.kualitas, ' - ', B.kualitasKet) AS [Kolektibilitas],
                B.jumlahHariTunggakan AS [Jumlah Hari Tunggakan],
                CONCAT(B.kodeSebabMacet, ' - ', B.sebabMacetKet) AS [Sebab Macet],
                B.tanggalMacet AS [Tanggal Macet],
                B.frekuensiTunggakan AS [Frekuensi Tunggakan],
                B.denda AS [Denda Biaya],
                B.frekuensiRestrukturisasi AS [Frekuensi Restrukturisasi],
                B.tanggalRestrukturisasiAkhir AS [Tanggal Restrukturisasi Terakhir]
            {base_query}
            ORDER BY [Nama Debitur]
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """

        # Eksekusi data query
        all_params = params + [offset, page_size]
        cursor.execute(data_query, all_params)
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        data = [dict(zip(columns, row)) for row in rows]

        # Buat string debug untuk ditampilkan di JSON response
        def build_debug_query(query, params):
            # Render param dengan tanda kutip jika string
            rendered_params = []
            for p in params:
                if isinstance(p, str):
                    rendered_params.append(f"'{p}'")
                else:
                    rendered_params.append(str(p))
            # Ganti tanda tanya (?) dengan parameter aktual
            full_query = query
            for rp in rendered_params:
                full_query = full_query.replace("?", rp, 1)
            return full_query.strip()

        return jsonify({
            'success': True,
            'data': data,
            'total': total_records,
            'page': page,
            'page_size': page_size,
            'debug': {
                'count_query': build_debug_query(count_query, params),
                'data_query': build_debug_query(data_query, all_params)
            }
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

    finally:
        cursor.close()
        conn.close()        

@app.route('/api/download-data-fasilitas-debitur', methods=['GET'])
def api_download_fasilitas_debitur_page():
    """
    Download data Excel berdasarkan filter periode (YYYYMM).
    """
    conn = None
    cursor = None
    try:
        periode = request.args.get('periode')
        if not periode:
            abort(400, description="Parameter 'periode' wajib diisi")

        conn = get_db_connection()
        cursor = conn.cursor()

        base_where = """
            WHERE A.bentukBU IS NOT NULL
              AND LEN(A.npwp) >= 5
              AND B.jenisKreditPembiayaan IS NOT NULL
        """
        periode_expr = f"LEFT(CONVERT(CHAR(8), CONVERT(DATE, '01 ' + {convert_bulan_sql('A.periodeData')}, 107), 112), 6)"
        base_where += f" AND {periode_expr} = ?"

        query = f"""
            SELECT
                CONVERT(CHAR(8),
                    EOMONTH(CONVERT(DATE, '01 ' + {convert_bulan_sql('A.periodeData')}, 107)), 112
                ) AS [Periode Data],
                A.npwp AS [ID Debitur/ NPWP],
                A.namaDebitur AS [Nama Debitur],
                CONCAT(B.jenisKreditPembiayaan, ' - ', B.jenisKreditPembiayaanKet) AS [Jenis Kredit / Pembiayaan],
                CONCAT(B.akadKreditPembiayaan, ' - ', B.akadKreditPembiayaanKet) AS [Akad Kredit / Pembiayaan],
                A.pelaporKet AS [Pelapor],
                B.bakiDebet AS [Baki Debet],
                CONCAT(B.kualitas, ' - ', B.kualitasKet) AS [Kolektibilitas],
                B.jumlahHariTunggakan AS [Jumlah Hari Tunggakan],
                CONCAT(B.kodeSebabMacet, ' - ', B.sebabMacetKet) AS [Sebab Macet],
                B.tanggalMacet AS [Tanggal Macet],
                B.frekuensiTunggakan AS [Frekuensi Tunggakan],
                B.denda AS [Denda Biaya],
                B.frekuensiRestrukturisasi AS [Frekuensi Restrukturisasi],
                B.tanggalRestrukturisasiAkhir AS [Tanggal Restrukturisasi Terakhir]
            FROM [dbo].[slik_data_pokok_debitur] A
            LEFT JOIN [dbo].[slik_fasilitas_kredit_pembiayaan] B
                ON A.namaFileUpload = B.namaFileUpload 
                AND A.periodeData = B.periodeData 
                AND A.pelaporKet = B.ljkKet
            {base_where}
            ORDER BY [Nama Debitur]
        """

        cursor.execute(query, [periode])
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]

        if not rows:
            abort(404, description="Tidak ada data untuk periode tersebut")

        # Buat workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fasilitas Debitur"

        # Header
        ws.append(columns)

        # Data
        for row in rows:
            ws.append(list(row))

        # Auto-width dan format angka
        numeric_cols = {"Baki Debet", "Jumlah Hari Tunggakan", "Denda Biaya", "Frekuensi Tunggakan", "Frekuensi Restrukturisasi"}
        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(r[col_idx - 1])) if r[col_idx - 1] is not None else 0 for r in rows)
            ws.column_dimensions[col_letter].width = min(max(max_len, len(col_name)) + 2, 50)
            if col_name in numeric_cols:
                for row_idx in range(2, len(rows) + 2):
                    ws.cell(row=row_idx, column=col_idx).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        # Simpan ke memory buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"fasilitas_debitur_{periode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# Data Kualitas Terburuk per Periode
@app.route('/data-kualitas-terburuk', methods=['GET'])
def data_kualitas_terburuk_page():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    return render_template(
        'data_kualitas_terburuk.html',
        username=session.get('username'),
        fullname=session.get('fullname'),
        role_access=session.get('role_access'),
        report_access=session.get('report_access')
    )

@app.route('/api/data-kualitas-terburuk', methods=['GET'])
def api_data_kualitas_terburuk_page():
    """
    Endpoint untuk data kualitas terburuk dengan filter periode (YYYYMM), pagination, dan limit.
    Menggunakan CTE optimized query (PeriodeCTE + FinalCTE).
    """
    conn = None
    cursor = None

    try:
        periode = request.args.get('periode')  # contoh: 202509
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 50))
        offset = (page - 1) * page_size

        conn = get_db_connection()
        cursor = conn.cursor()

        # ============================
        # PARAMETER
        # ============================
        params = []
        periode_filter = ""

        if periode:
            periode_filter = f"""WHERE LEFT(P.periode_yyyymmdd, 6) = ?
            AND B.kualitasTerburuk IS NOT NULL
            AND B.kualitasBulanDataTerburuk IS NOT NULL"""
            params.append(periode)

        # ============================
        # QUERY HITUNG TOTAL
        # ============================
        count_query = f"""
            WITH PeriodeCTE AS (
                SELECT 
                    A.periodeData,
                    A.namaDebitur,
                    A.npwp,
                    A.namaFileUpload,
                    A.uploadDate,
                    FORMAT(
                        EOMONTH(
                            DATEFROMPARTS(
                                CAST(RIGHT(A.periodeData, 4) AS INT),
                                CASE LEFT(A.periodeData, CHARINDEX(' ', A.periodeData) - 1)
                                    WHEN 'Januari'   THEN 1
                                    WHEN 'Februari'  THEN 2
                                    WHEN 'Maret'     THEN 3
                                    WHEN 'April'     THEN 4
                                    WHEN 'Mei'       THEN 5
                                    WHEN 'Juni'      THEN 6
                                    WHEN 'Juli'      THEN 7
                                    WHEN 'Agustus'   THEN 8
                                    WHEN 'September' THEN 9
                                    WHEN 'Oktober'   THEN 10
                                    WHEN 'November'  THEN 11
                                    WHEN 'Desember'  THEN 12
                                END,
                                1
                            )
                        ),
                        'yyyyMMdd'
                    ) AS periode_yyyymmdd
                FROM slik_data_pokok_debitur A
                WHERE A.bentukBU IS NOT NULL AND A.npwp IS NOT NULL
            ),
            FinalCTE AS (
                SELECT
                    LEFT(P.periode_yyyymmdd, 6) AS periode_yyyymm,
                    P.namaDebitur,
                    P.npwp,
                    B.kualitasTerburuk,
                    B.kualitasBulanDataTerburuk,
                    ROW_NUMBER() OVER (
                        PARTITION BY LEFT(P.periode_yyyymmdd, 6), P.namaDebitur, P.npwp
                        ORDER BY P.uploadDate DESC, P.namaFileUpload ASC
                    ) AS rn
                FROM PeriodeCTE P
                INNER JOIN slik_ringkasan_fasilitas B
                    ON P.namaFileUpload = B.namaFileUpload
                    AND P.uploadDate = B.uploadDate
                    AND P.periodeData = B.periodeData
                    AND B.kualitasTerburuk <> ''
                    AND B.kualitasBulanDataTerburuk <> ''
                {periode_filter}
            )
            SELECT COUNT(*) FROM FinalCTE WHERE rn = 1;
        """

        cursor.execute(count_query, params)
        total_records = cursor.fetchone()[0]

        # ============================
        # QUERY DATA PAGINATED
        # ============================
        data_query = f"""
            WITH PeriodeCTE AS (
                SELECT 
                    A.periodeData,
                    A.namaDebitur,
                    A.npwp,
                    A.namaFileUpload,
                    A.uploadDate,
                    FORMAT(
                        EOMONTH(
                            DATEFROMPARTS(
                                CAST(RIGHT(A.periodeData, 4) AS INT),
                                CASE LEFT(A.periodeData, CHARINDEX(' ', A.periodeData) - 1)
                                    WHEN 'Januari'   THEN 1
                                    WHEN 'Februari'  THEN 2
                                    WHEN 'Maret'     THEN 3
                                    WHEN 'April'     THEN 4
                                    WHEN 'Mei'       THEN 5
                                    WHEN 'Juni'      THEN 6
                                    WHEN 'Juli'      THEN 7
                                    WHEN 'Agustus'   THEN 8
                                    WHEN 'September' THEN 9
                                    WHEN 'Oktober'   THEN 10
                                    WHEN 'November'  THEN 11
                                    WHEN 'Desember'  THEN 12
                                END,
                                1
                            )
                        ),
                        'yyyyMMdd'
                    ) AS periode_yyyymmdd
                FROM slik_data_pokok_debitur A
                WHERE A.bentukBU IS NOT NULL AND A.npwp IS NOT NULL
            ),
            FinalCTE AS (
                SELECT
                    LEFT(P.periode_yyyymmdd, 6) AS periode_yyyymm,
                    P.namaDebitur,
                    P.npwp,
                    B.kualitasTerburuk,
                    B.kualitasBulanDataTerburuk,
                    ROW_NUMBER() OVER (
                        PARTITION BY LEFT(P.periode_yyyymmdd, 6), P.namaDebitur, P.npwp
                        ORDER BY P.uploadDate DESC, P.namaFileUpload ASC
                    ) AS rn
                FROM PeriodeCTE P
                INNER JOIN slik_ringkasan_fasilitas B
                    ON P.namaFileUpload = B.namaFileUpload
                    AND P.uploadDate = B.uploadDate
                    AND P.periodeData = B.periodeData
                    AND B.kualitasTerburuk <> ''
                    AND B.kualitasBulanDataTerburuk <> ''
                {periode_filter}
            )
            SELECT
                periode_yyyymm AS [Periode Data],
                namaDebitur AS [Nama Debitur],
                npwp AS [ID Debitur NPWP],
                kualitasTerburuk AS [Kualitas Terburuk],
                kualitasBulanDataTerburuk AS [Kualitas Bulan Data Terburuk]
            FROM FinalCTE
            WHERE rn = 1
            ORDER BY namaDebitur
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY;
        """

        cursor.execute(data_query, params + [offset, page_size])
        rows = cursor.fetchall()
        columns = [c[0] for c in cursor.description]
        data = [dict(zip(columns, r)) for r in rows]

        return jsonify({
            "success": True,
            "total": total_records,
            "page": page,
            "page_size": page_size,
            "data": data
        })

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
        
@app.route('/api/download-data-kualitas-terburuk', methods=['GET'])
def api_download_data_kualitas_terburuk_page():
    """
    Download data kualitas terburuk (Excel) berdasarkan filter periode (YYYYMM).
    Menggunakan CTE optimized query (PeriodeCTE + FinalCTE).
    """
    conn = None
    cursor = None

    try:
        periode = request.args.get('periode')   # contoh: 202509

        conn = get_db_connection()
        cursor = conn.cursor()

        params = []
        periode_filter = ""

        if periode:
            periode_filter = f"""WHERE LEFT(P.periode_yyyymmdd, 6) = ?
            AND B.kualitasTerburuk IS NOT NULL
            AND B.kualitasBulanDataTerburuk IS NOT NULL"""
            params.append(periode)

        # ============================
        # QUERY DOWNLOAD FULL DATA
        # ============================
        download_query = f"""
            WITH PeriodeCTE AS (
                SELECT 
                    A.periodeData,
                    A.namaDebitur,
                    A.npwp,
                    A.namaFileUpload,
                    A.uploadDate,
                    FORMAT(
                        EOMONTH(
                            DATEFROMPARTS(
                                CAST(RIGHT(A.periodeData, 4) AS INT),
                                CASE LEFT(A.periodeData, CHARINDEX(' ', A.periodeData) - 1)
                                    WHEN 'Januari'   THEN 1
                                    WHEN 'Februari'  THEN 2
                                    WHEN 'Maret'     THEN 3
                                    WHEN 'April'     THEN 4
                                    WHEN 'Mei'       THEN 5
                                    WHEN 'Juni'      THEN 6
                                    WHEN 'Juli'      THEN 7
                                    WHEN 'Agustus'   THEN 8
                                    WHEN 'September' THEN 9
                                    WHEN 'Oktober'   THEN 10
                                    WHEN 'November'  THEN 11
                                    WHEN 'Desember'  THEN 12
                                END,
                                1
                            )
                        ),
                        'yyyyMMdd'
                    ) AS periode_yyyymmdd
                FROM slik_data_pokok_debitur A
                WHERE A.bentukBU IS NOT NULL AND A.npwp IS NOT NULL
            ),
            FinalCTE AS (
                SELECT
                    LEFT(P.periode_yyyymmdd, 6) AS periode_yyyymm,
                    P.namaDebitur,
                    P.npwp,
                    B.kualitasTerburuk,
                    B.kualitasBulanDataTerburuk,
                    ROW_NUMBER() OVER (
                        PARTITION BY LEFT(P.periode_yyyymmdd, 6), P.namaDebitur, P.npwp
                        ORDER BY P.uploadDate DESC, P.namaFileUpload ASC
                    ) AS rn
                FROM PeriodeCTE P
                INNER JOIN slik_ringkasan_fasilitas B
                    ON P.namaFileUpload = B.namaFileUpload
                    AND P.uploadDate = B.uploadDate
                    AND P.periodeData = B.periodeData
                    AND B.kualitasTerburuk <> ''
                    AND B.kualitasBulanDataTerburuk <> ''
                {periode_filter}
            )
            SELECT
                periode_yyyymm AS [Periode Data],
                namaDebitur AS [Nama Debitur],
                npwp AS [ID Debitur NPWP],
                kualitasTerburuk AS [Kualitas Terburuk],
                kualitasBulanDataTerburuk AS [Kualitas Bulan Data Terburuk]
            FROM FinalCTE
            WHERE rn = 1
            ORDER BY namaDebitur;
        """

        cursor.execute(download_query, params)
        rows = cursor.fetchall()
        columns = [c[0] for c in cursor.description]

        if not rows:
            return jsonify({
                "success": False,
                "message": "Tidak ada data untuk periode tersebut."
            })

        # ============================
        # EXPORT TO EXCEL
        # ============================
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Kualitas Terburuk"

        # Header
        ws.append(columns)

        # Data rows
        for row in rows:
            ws.append(list(row))

        # Auto column width
        for i, col in enumerate(columns, start=1):
            max_len = max(len(str(r[i-1])) if r[i-1] else 0 for r in rows)
            ws.column_dimensions[get_column_letter(i)].width = min(max(max_len, len(col)) + 2, 50)

        # Output stream
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"data_kualitas_terburuk_{periode or 'all'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

    finally:
        if cursor: cursor.close()
        if conn: conn.close()

def get_tax_cte() -> str:
    """
    CTE gabungan Debitur Aktif (Konvensional + Syariah)
    Menggunakan parameter @periode (YYYYMM)
    Query mengikuti versi SQL terbaru dari user.
    """
    return """
    ;WITH TAX AS
    (
        -----------------------------------------
        --             KONVENSIONAL
        -----------------------------------------
        SELECT
            A.PBK_EOD_DATE,
            A.CUSTOMER_ID,
            A.CUST_NAME,
            D02.Nomor_Identitas_Badan_Usaha,

            CASE
                WHEN SUM(CASE WHEN F.FACILITY_STATUS = 'AC' THEN 1 ELSE 0 END) > 0
                    THEN 'Active'
                ELSE 'Closed'
            END AS CUSTOMER_STATUS

        FROM [10.10.4.12].SMIDWHARIUM.dbo.PBK_M_CIF_CUSTOMER A

        LEFT JOIN (
            SELECT *
            FROM [10.10.4.12].SMIDWHARIUM.dbo.PBK_T_INF_COR_FACILITY_ACCOUNT
            WHERE PRODUCT_ID NOT IN (40,1083,1082,1084,1086)
        ) F
            ON A.CUSTOMER_ID = F.CUSTOMER_ID
           AND A.PBK_EOD_DATE = F.PBK_EOD_DATE

        LEFT JOIN [10.10.4.12].SMIDWH_REPSLIK.dbo.FORM_D02 D02
            ON YEAR(A.PBK_EOD_DATE) = YEAR(D02.Tanggal_Data)
           AND MONTH(A.PBK_EOD_DATE) = MONTH(D02.Tanggal_Data)
           AND (
                CASE
                    WHEN A.CUSTOMER_ID IN ('433','404') THEN '711'   -- Angkasa Pura
                    WHEN A.CUSTOMER_ID = '697' THEN '699'           -- TJT
                    ELSE A.CUSTOMER_ID
                END
           ) = D02.[Nomor_CIF Debitur]

        WHERE
            A.PBK_EOD_DATE = (
                SELECT MAX(PBK_EOD_DATE)
                FROM [10.10.4.12].SMIDWHARIUM.dbo.PBK_M_CIF_CUSTOMER
                WHERE @periode IS NULL
                   OR CONVERT(CHAR(6), PBK_EOD_DATE, 112) = @periode
            )
            AND F.FACILITY_NO NOT IN (SELECT ACC_NO FROM [10.10.4.12].SMIDWHARIUM.dbo.PBK_DEBTOR_EXCLUDE)
            AND A.PBK_EOD_DATE <= (
                SELECT MAX(Tanggal_Data)
                FROM [10.10.4.12].SMIDWH_REPSLIK.dbo.FORM_D02
            )

        GROUP BY
            A.PBK_EOD_DATE,
            A.CUSTOMER_ID,
            A.CUST_NAME,
            D02.Nomor_Identitas_Badan_Usaha


        -----------------------------------------
        UNION ALL
        -----------------------------------------

        -----------------------------------------
        --               SYARIAH
        -----------------------------------------
        SELECT
            A.[EOD DATE] AS PBK_EOD_DATE,
            A.ID AS CUSTOMER_ID,
            A.[NAME 1] AS CUST_NAME,
            D02.Nomor_Identitas_Badan_Usaha,
            'Active' AS CUSTOMER_STATUS

        FROM [10.10.4.12].SMIDWHTEMENOS.dbo.T24_CUSTOMER A

        LEFT JOIN [10.10.4.12].SMIDWH_REPSLIK.dbo.FORM_D02 D02
            ON YEAR(A.[EOD DATE]) = YEAR(D02.Tanggal_Data)
           AND MONTH(A.[EOD DATE]) = MONTH(D02.Tanggal_Data)
           AND A.ID = D02.[Nomor_CIF Debitur]

        LEFT JOIN (
            SELECT DISTINCT [CUSTOMER ID], STATUS, [EOD DATE]
            FROM [10.10.4.12].SMIDWHTEMENOS.dbo.T24_LD_LOANS_AND_DEPOSITS
        ) B
            ON A.ID = B.[CUSTOMER ID]
           AND A.[EOD DATE] = B.[EOD DATE]

        WHERE
            A.[EOD DATE] = (
                SELECT MAX([EOD DATE])
                FROM [10.10.4.12].SMIDWHTEMENOS.dbo.T24_CUSTOMER
                WHERE @periode IS NULL
                   OR CONVERT(CHAR(6), [EOD DATE], 112) = @periode
            )
            AND A.[EOD DATE] <= (
                SELECT MAX(Tanggal_Data)
                FROM [10.10.4.12].SMIDWH_REPSLIK.dbo.FORM_D02
            )
            AND STATUS = 'CUR'
    )
    """

@app.route('/api/maxslik', methods=['GET'])
def api_maxslik():
    try:
        with get_db_connection() as conn, conn.cursor() as cursor:
            cursor.execute("SELECT CONVERT(CHAR(6), MAX(Tanggal_Data), 112) FROM [10.10.4.12].SMIDWH_REPSLIK.dbo.FORM_D02")
            maxslik = cursor.fetchone()[0]
        return jsonify({'success': True, 'maxslik': maxslik})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Data NPWP Debitur Aktif per Periode
@app.route('/data-debitur-aktif', methods=['GET'])
def data_debitur_aktif_page():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    return render_template(
        'data_debitur_aktif.html',
        username=session.get('username'),
        fullname=session.get('fullname'),
        role_access=session.get('role_access'),
        report_access=session.get('report_access')
    )
            
@app.route('/api/data-debitur-aktif', methods=['GET'])
def api_data_debitur_aktif_page():
    try:
        periode = request.args.get('periode')  # contoh: 202510
        page = max(int(request.args.get('page', 1)), 1)
        page_size = min(int(request.args.get('page_size', 50)), 1000)
        offset = (page - 1) * page_size

        # ==========================
        # Ambil MAXSLIK dari FORM_D02
        # ==========================
        with get_db_connection() as conn, conn.cursor() as cursor:
            cursor.execute("SELECT CONVERT(CHAR(6), MAX(Tanggal_Data), 112) FROM [10.10.4.12].SMIDWH_REPSLIK.dbo.FORM_D02")
            maxslik = cursor.fetchone()[0]

        # Jika periode > maxslik → langsung balikan 0 data
        if periode and periode > maxslik:
            return jsonify({
                'success': True,
                'data': [],
                'total': 0,
                'page': page,
                'page_size': page_size,
                'reason': 'periode_gt_maxslik'
            })

        # ==========================
        # Build CTE query
        # ==========================
        cte_query = get_tax_cte()

        # Query COUNT
        query_count = f"""
            DECLARE @periode CHAR(6) = ?;
            {cte_query}
            SELECT COUNT(*) AS total
            FROM TAX
            WHERE CUSTOMER_STATUS = 'Active';
        """

        # Query Data
        query_data = f"""
            DECLARE @periode CHAR(6) = ?;
            {cte_query}
            SELECT *
            FROM TAX
            WHERE CUSTOMER_STATUS = 'Active'
            ORDER BY CUSTOMER_ID
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY;
        """

        # Eksekusi query
        with get_db_connection() as conn, conn.cursor() as cursor:
            cursor.execute(query_count, [periode])
            total = cursor.fetchone()[0]

            cursor.execute(query_data, [periode, offset, page_size])
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in cursor.fetchall()]

        return jsonify({
            'success': True,
            'data': data,
            'total': total,
            'page': page,
            'page_size': page_size,
            'maxslik': maxslik
        })

    except Exception as e:
        logger.exception("Error pada api_data_debitur_aktif_page")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/download-data-debitur-aktif', methods=['POST'])
def api_download_data_debitur_aktif_page():
    """
    Download data Debitur Aktif ke Excel (identik dengan hasil /api/data-debitur-aktif)
    """
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first.'})

    try:
        periode = request.json.get('periode')  # contoh: "202511" atau None

        # ==========================
        # Ambil MAXSLIK
        # ==========================
        with get_db_connection() as conn, conn.cursor() as cursor:
            cursor.execute("SELECT CONVERT(CHAR(6), MAX(Tanggal_Data), 112) FROM SMIDWH_REPSLIK..FORM_D02")
            maxslik = cursor.fetchone()[0]

        # Jika periode > maxslik → tidak boleh download
        if periode and periode > maxslik:
            return jsonify({
                'success': False,
                'message': f'Tidak ada data untuk periode di atas MAXSLIK ({maxslik}).'
            })

        # ==========================
        # Query data utama
        # ==========================
        query = f"""
            DECLARE @periode CHAR(6) = ?;
            {get_tax_cte()}
            SELECT *
            FROM TAX
            WHERE CUSTOMER_STATUS = 'Active'
            ORDER BY CUSTOMER_ID;
        """

        with get_db_connection() as conn, conn.cursor() as cursor:
            cursor.execute(query, [periode])
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]

        # ==========================
        # Buat file Excel
        # ==========================
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Debitur Aktif"
        ws.append(columns)
        ws.freeze_panes = "A2"

        thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Isi data
        for row in rows:
            ws.append(list(row))

        # Format kolom otomatis
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = len(col_name)
            for row_idx in range(2, len(rows) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin
                if isinstance(cell.value, (int, float)):
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        # Simpan stream
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"data_debitur_aktif_{periode or 'all'}.xlsx"
        logger.info(f"User {session['username']} download data periode {periode}")

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.exception("Error pada api_download_data_debitur_aktif_page")
        return jsonify({'success': False, 'message': str(e)})

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, threaded=True, debug=True)
